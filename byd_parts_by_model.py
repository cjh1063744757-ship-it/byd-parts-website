"""
BYD Parts Catalog - Organized by Vehicle Model
CM IMPORT&EXPORT
Output: BYD_Parts_By_Model_CM_EXPORT.xlsx
"""

import openpyxl
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
import os

OUTPUT_FILE = r'C:\Users\Administrator\Desktop\BYD_Parts_By_Model_CM_EXPORT.xlsx'

# ─── Colour palette ──────────────────────────────────────────────────────────
C_BYD_RED    = "C0392B"   # BYD brand red
C_DARK       = "1A1A2E"   # dark navy header
C_MID        = "16213E"
C_ACCENT     = "0F3460"
C_GOLD       = "E94560"
C_LIGHT_GRAY = "F2F2F2"
C_WHITE      = "FFFFFF"

# Category colours (one per system category)
CAT_COLORS = {
    "Body & Exterior"       : "2E86AB",
    "Lighting"              : "F6AE2D",
    "Doors, Glass & Mirrors": "F26419",
    "Brakes"                : "C0392B",
    "Suspension & Steering" : "27AE60",
    "EV Drivetrain"         : "8E44AD",
    "Cooling System"        : "2980B9",
    "Interior"              : "7F8C8D",
    "Wheels & Tyres"        : "E67E22",
    "Maintenance"           : "16A085",
}

# ─── Model data ──────────────────────────────────────────────────────────────
# Tuple: (category, name_en, name_cn, part_no, price_usd, moq, condition, notes)

MODELS = {

# ═══════════════════════════════════════════════════════════════════════════════
"BYD E2": {
    "info": {"year":"2019–2024","type":"Compact EV Hatchback","voltage":"60V Li-Fe","engine":"TZ180XSY 90kW"},
    "parts": [
        # Body & Exterior
        ("Body & Exterior","Front Bumper Assembly","前保险杠总成","E2-BDY-001",38.00,5,"OEM New","With fog lamp hole"),
        ("Body & Exterior","Rear Bumper Assembly","后保险杠总成","E2-BDY-002",35.00,5,"OEM New",""),
        ("Body & Exterior","Hood / Bonnet","发动机盖","E2-BDY-003",65.00,3,"OEM New","Steel"),
        ("Body & Exterior","Front Left Fender","左前翼子板","E2-BDY-004",28.00,4,"OEM New",""),
        ("Body & Exterior","Front Right Fender","右前翼子板","E2-BDY-005",28.00,4,"OEM New",""),
        ("Body & Exterior","Trunk Lid","行李箱盖","E2-BDY-006",72.00,3,"OEM New",""),
        ("Body & Exterior","Side Skirt Set","侧裙总成","E2-BDY-007",22.00,5,"OEM New","Pair"),
        # Lighting
        ("Lighting","Left Headlight Assembly","左前大灯总成","E2-LGT-001",58.00,2,"OEM New","LED DRL incl."),
        ("Lighting","Right Headlight Assembly","右前大灯总成","E2-LGT-002",58.00,2,"OEM New","LED DRL incl."),
        ("Lighting","Left Tail Light Assembly","左后尾灯总成","E2-LGT-003",42.00,2,"OEM New",""),
        ("Lighting","Right Tail Light Assembly","右后尾灯总成","E2-LGT-004",42.00,2,"OEM New",""),
        ("Lighting","Front Left Fog Lamp","左前雾灯","E2-LGT-005",12.00,4,"OEM New",""),
        ("Lighting","Front Right Fog Lamp","右前雾灯","E2-LGT-006",12.00,4,"OEM New",""),
        ("Lighting","High-Mount Stop Lamp","高位刹车灯","E2-LGT-007",8.50,6,"OEM New",""),
        # Doors, Glass & Mirrors
        ("Doors, Glass & Mirrors","Front Left Door Assembly","左前门总成","E2-DGM-001",118.00,2,"OEM New","Bare shell"),
        ("Doors, Glass & Mirrors","Front Right Door Assembly","右前门总成","E2-DGM-002",118.00,2,"OEM New","Bare shell"),
        ("Doors, Glass & Mirrors","Rear Left Door Assembly","左后门总成","E2-DGM-003",112.00,2,"OEM New",""),
        ("Doors, Glass & Mirrors","Rear Right Door Assembly","右后门总成","E2-DGM-004",112.00,2,"OEM New",""),
        ("Doors, Glass & Mirrors","Windshield (Front)","前挡风玻璃","E2-DGM-005",55.00,2,"OEM New","Laminated"),
        ("Doors, Glass & Mirrors","Rear Window Glass","后挡风玻璃","E2-DGM-006",38.00,2,"OEM New","Heated"),
        ("Doors, Glass & Mirrors","Left Side Mirror Assy","左后视镜总成","E2-DGM-007",32.00,2,"OEM New","Electric fold"),
        ("Doors, Glass & Mirrors","Right Side Mirror Assy","右后视镜总成","E2-DGM-008",32.00,2,"OEM New","Electric fold"),
        ("Doors, Glass & Mirrors","Door Weatherstrip Set","门密封条套装","E2-DGM-009",18.00,5,"OEM New","4pcs set"),
        # Brakes
        ("Brakes","Front Brake Pad Set","前刹车片套装","E2-BRK-001",14.50,10,"OEM New","Ceramic"),
        ("Brakes","Rear Brake Pad Set","后刹车片套装","E2-BRK-002",13.00,10,"OEM New",""),
        ("Brakes","Front Brake Disc","前刹车盘","E2-BRK-003",22.00,6,"OEM New","Vented"),
        ("Brakes","Rear Brake Disc","后刹车盘","E2-BRK-004",19.00,6,"OEM New","Solid"),
        ("Brakes","Brake Master Cylinder","制动主缸","E2-BRK-005",38.00,3,"OEM New",""),
        # Suspension & Steering
        ("Suspension & Steering","Front Shock Absorber (L)","左前减振器","E2-SUS-001",35.00,2,"OEM New",""),
        ("Suspension & Steering","Front Shock Absorber (R)","右前减振器","E2-SUS-002",35.00,2,"OEM New",""),
        ("Suspension & Steering","Rear Shock Absorber (L)","左后减振器","E2-SUS-003",30.00,2,"OEM New",""),
        ("Suspension & Steering","Rear Shock Absorber (R)","右后减振器","E2-SUS-004",30.00,2,"OEM New",""),
        ("Suspension & Steering","Front Lower Control Arm L","左前下摆臂","E2-SUS-005",28.00,4,"OEM New",""),
        ("Suspension & Steering","Front Wheel Bearing L","左前轮毂轴承","E2-SUS-006",16.00,4,"OEM New",""),
        ("Suspension & Steering","Tie Rod End L","左横拉杆球头","E2-SUS-007",9.50,6,"OEM New",""),
        # EV Drivetrain
        ("EV Drivetrain","OBC On-Board Charger","车载充电机","E2-EV-001",185.00,1,"OEM New","6.6kW AC"),
        ("EV Drivetrain","DC-DC Converter","DC-DC转换器","E2-EV-002",95.00,2,"OEM New","12V output"),
        ("EV Drivetrain","BMS Battery Management","电池管理系统","E2-EV-003",220.00,1,"OEM New",""),
        ("EV Drivetrain","Charging Port Assembly","充电口总成","E2-EV-004",28.00,5,"OEM New","CCS2 / CHAdeMO"),
        ("EV Drivetrain","12V Auxiliary Battery","12V辅助电池","E2-EV-005",42.00,3,"OEM New","40Ah"),
        ("EV Drivetrain","Motor Controller (MCU)","电机控制器","E2-EV-006",320.00,1,"OEM New","90kW"),
        # Cooling
        ("Cooling System","Radiator Assembly","水箱总成","E2-COL-001",52.00,3,"OEM New",""),
        ("Cooling System","Battery Cooling Pump","电池冷却水泵","E2-COL-002",38.00,3,"OEM New",""),
        ("Cooling System","HVAC Compressor","空调压缩机","E2-COL-003",118.00,2,"OEM New","EV electric type"),
        # Interior
        ("Interior","Floor Mat Set","脚垫套装","E2-INT-001",18.00,10,"OEM New","4pcs"),
        ("Interior","Seat Cover Set","座椅套","E2-INT-002",35.00,5,"Aftermarket","Full set"),
        ("Interior","Dashboard Assembly","仪表板总成","E2-INT-003",145.00,1,"OEM New",""),
        # Wheels & Tyres
        ("Wheels & Tyres","Alloy Wheel 15in","15寸铝合金轮毂","E2-WHL-001",38.00,4,"OEM New","5×100 ET40"),
        ("Wheels & Tyres","Tyre 185/60R15","轮胎 185/60R15","E2-WHL-002",32.00,4,"OEM New","Recommend EV spec"),
        # Maintenance
        ("Maintenance","Cabin Air Filter","空调滤芯","E2-MNT-001",6.50,20,"OEM New",""),
        ("Maintenance","Wiper Blade Set","雨刮器套装","E2-MNT-002",8.00,10,"OEM New","Front pair"),
        ("Maintenance","Brake Fluid DOT4","制动液DOT4","E2-MNT-003",5.00,20,"OEM New","500ml"),
    ]
},

# ═══════════════════════════════════════════════════════════════════════════════
"BYD Atto 3": {
    "info": {"year":"2021–2024","type":"Compact EV SUV","voltage":"400V Li-Fe / NMC","engine":"EF7A 150kW"},
    "parts": [
        ("Body & Exterior","Front Bumper Assembly","前保险杠总成","AT3-BDY-001",48.00,4,"OEM New","Sport trim"),
        ("Body & Exterior","Rear Bumper Assembly","后保险杠总成","AT3-BDY-002",45.00,4,"OEM New",""),
        ("Body & Exterior","Hood / Bonnet","发动机盖","AT3-BDY-003",78.00,2,"OEM New","Steel"),
        ("Body & Exterior","Front Left Fender","左前翼子板","AT3-BDY-004",35.00,4,"OEM New",""),
        ("Body & Exterior","Front Right Fender","右前翼子板","AT3-BDY-005",35.00,4,"OEM New",""),
        ("Body & Exterior","Trunk Lid","行李箱盖","AT3-BDY-006",88.00,2,"OEM New","With spoiler"),
        ("Body & Exterior","Roof Rails Set","车顶行李架","AT3-BDY-007",42.00,5,"OEM New","Pair"),
        ("Lighting","Left Headlight Assembly","左前大灯总成","AT3-LGT-001",78.00,2,"OEM New","Full LED matrix"),
        ("Lighting","Right Headlight Assembly","右前大灯总成","AT3-LGT-002",78.00,2,"OEM New","Full LED matrix"),
        ("Lighting","Left Tail Light Assembly","左后尾灯总成","AT3-LGT-003",55.00,2,"OEM New","LED sequential"),
        ("Lighting","Right Tail Light Assembly","右后尾灯总成","AT3-LGT-004",55.00,2,"OEM New","LED sequential"),
        ("Lighting","DRL Strip Left","左日间行车灯","AT3-LGT-005",18.00,4,"OEM New",""),
        ("Lighting","DRL Strip Right","右日间行车灯","AT3-LGT-006",18.00,4,"OEM New",""),
        ("Doors, Glass & Mirrors","Front Left Door Assembly","左前门总成","AT3-DGM-001",135.00,2,"OEM New",""),
        ("Doors, Glass & Mirrors","Front Right Door Assembly","右前门总成","AT3-DGM-002",135.00,2,"OEM New",""),
        ("Doors, Glass & Mirrors","Rear Left Door Assembly","左后门总成","AT3-DGM-003",128.00,2,"OEM New",""),
        ("Doors, Glass & Mirrors","Rear Right Door Assembly","右后门总成","AT3-DGM-004",128.00,2,"OEM New",""),
        ("Doors, Glass & Mirrors","Windshield (Front)","前挡风玻璃","AT3-DGM-005",68.00,2,"OEM New","Acoustic laminated"),
        ("Doors, Glass & Mirrors","Left Side Mirror Assy","左后视镜总成","AT3-DGM-006",42.00,2,"OEM New","Blind spot sensor"),
        ("Doors, Glass & Mirrors","Right Side Mirror Assy","右后视镜总成","AT3-DGM-007",42.00,2,"OEM New","Blind spot sensor"),
        ("Doors, Glass & Mirrors","Door Handle Set","门把手套装","AT3-DGM-008",28.00,5,"OEM New","4pcs, flush type"),
        ("Brakes","Front Brake Pad Set","前刹车片套装","AT3-BRK-001",17.50,10,"OEM New",""),
        ("Brakes","Rear Brake Pad Set","后刹车片套装","AT3-BRK-002",15.50,10,"OEM New",""),
        ("Brakes","Front Brake Disc","前刹车盘","AT3-BRK-003",28.00,6,"OEM New","Vented 296mm"),
        ("Brakes","Rear Brake Disc","后刹车盘","AT3-BRK-004",24.00,6,"OEM New",""),
        ("Brakes","EPB Actuator Rear","后电子驻车执行器","AT3-BRK-005",58.00,2,"OEM New",""),
        ("Suspension & Steering","Front Shock Absorber (L)","左前减振器","AT3-SUS-001",45.00,2,"OEM New",""),
        ("Suspension & Steering","Front Shock Absorber (R)","右前减振器","AT3-SUS-002",45.00,2,"OEM New",""),
        ("Suspension & Steering","Rear Shock Absorber (L)","左后减振器","AT3-SUS-003",40.00,2,"OEM New",""),
        ("Suspension & Steering","Rear Shock Absorber (R)","右后减振器","AT3-SUS-004",40.00,2,"OEM New",""),
        ("Suspension & Steering","Front Subframe","前副车架","AT3-SUS-005",125.00,1,"OEM New",""),
        ("Suspension & Steering","Steering Rack Assembly","转向机总成","AT3-SUS-006",185.00,1,"OEM New","EPS"),
        ("EV Drivetrain","OBC On-Board Charger 11kW","车载充电机11kW","AT3-EV-001",245.00,1,"OEM New","Three-phase AC"),
        ("EV Drivetrain","DC-DC Converter","DC-DC转换器","AT3-EV-002",115.00,1,"OEM New","14V 3kW"),
        ("EV Drivetrain","BMS Battery Management","电池管理系统","AT3-EV-003",280.00,1,"OEM New","Blade battery"),
        ("EV Drivetrain","Charging Port CCS2","充电口CCS2","AT3-EV-004",35.00,4,"OEM New","DC fast charge"),
        ("EV Drivetrain","Front Drive Motor 150kW","前驱电机150kW","AT3-EV-005",680.00,1,"OEM New","Permanent magnet"),
        ("EV Drivetrain","Motor Controller (MCU)","电机控制器","AT3-EV-006",420.00,1,"OEM New",""),
        ("Cooling System","Battery Thermal Management","电池热管理总成","AT3-COL-001",185.00,1,"OEM New","Heat pump type"),
        ("Cooling System","HVAC Compressor Electric","电动空调压缩机","AT3-COL-002",148.00,1,"OEM New",""),
        ("Cooling System","Coolant Expansion Tank","冷却液膨胀壶","AT3-COL-003",18.00,6,"OEM New",""),
        ("Interior","Steering Wheel Assembly","方向盘总成","AT3-INT-001",88.00,2,"OEM New","Heated, multi-func"),
        ("Interior","Center Console Assembly","中控台总成","AT3-INT-002",220.00,1,"OEM New","12.8in Pad base"),
        ("Interior","Floor Mat Set Premium","豪华脚垫套装","AT3-INT-003",28.00,8,"OEM New","3D TPE 5pcs"),
        ("Wheels & Tyres","Alloy Wheel 18in","18寸铝合金轮毂","AT3-WHL-001",62.00,4,"OEM New","5×114.3 ET40"),
        ("Wheels & Tyres","Tyre 215/55R18","轮胎 215/55R18","AT3-WHL-002",52.00,4,"OEM New","EV spec low rolling"),
        ("Maintenance","Cabin Air Filter w/ PM2.5","PM2.5空调滤芯","AT3-MNT-001",9.50,15,"OEM New",""),
        ("Maintenance","Wiper Blade Set","雨刮器套装","AT3-MNT-002",10.00,10,"OEM New","Flat blade pair"),
    ]
},

# ═══════════════════════════════════════════════════════════════════════════════
"BYD Seal": {
    "info": {"year":"2022–2024","type":"Mid-size EV Sedan","voltage":"800V CTB","engine":"EF7A-HS 390kW AWD"},
    "parts": [
        ("Body & Exterior","Front Bumper Assembly","前保险杠总成","SEA-BDY-001",62.00,3,"OEM New","Performance trim"),
        ("Body & Exterior","Rear Bumper Assembly","后保险杠总成","SEA-BDY-002",58.00,3,"OEM New","Diffuser style"),
        ("Body & Exterior","Hood / Bonnet","发动机盖","SEA-BDY-003",95.00,2,"OEM New","Steel bonded"),
        ("Body & Exterior","Front Left Fender","左前翼子板","SEA-BDY-004",42.00,3,"OEM New",""),
        ("Body & Exterior","Front Right Fender","右前翼子板","SEA-BDY-005",42.00,3,"OEM New",""),
        ("Body & Exterior","Trunk Lid","行李箱盖","SEA-BDY-006",105.00,2,"OEM New","Motorized"),
        ("Body & Exterior","Frunk Lid (Front Trunk)","前备厢盖","SEA-BDY-007",68.00,2,"OEM New","AWD models"),
        ("Lighting","Left Full-LED Headlight","左全LED大灯","SEA-LGT-001",115.00,2,"OEM New","Matrix ADB"),
        ("Lighting","Right Full-LED Headlight","右全LED大灯","SEA-LGT-002",115.00,2,"OEM New","Matrix ADB"),
        ("Lighting","Left Tail Light Assembly","左后尾灯总成","SEA-LGT-003",72.00,2,"OEM New","Full LED bar"),
        ("Lighting","Right Tail Light Assembly","右后尾灯总成","SEA-LGT-004",72.00,2,"OEM New","Full LED bar"),
        ("Lighting","Interior Ambient Light Kit","内饰氛围灯套件","SEA-LGT-005",35.00,5,"OEM New","64-color"),
        ("Doors, Glass & Mirrors","Front Left Door Assembly","左前门总成","SEA-DGM-001",158.00,2,"OEM New","Frameless window"),
        ("Doors, Glass & Mirrors","Front Right Door Assembly","右前门总成","SEA-DGM-002",158.00,2,"OEM New",""),
        ("Doors, Glass & Mirrors","Rear Left Door Assembly","左后门总成","SEA-DGM-003",148.00,2,"OEM New",""),
        ("Doors, Glass & Mirrors","Rear Right Door Assembly","右后门总成","SEA-DGM-004",148.00,2,"OEM New",""),
        ("Doors, Glass & Mirrors","Panoramic Glass Roof","全景天窗玻璃","SEA-DGM-005",195.00,1,"OEM New","Large fixed glass"),
        ("Doors, Glass & Mirrors","Windshield HUD","HUD前挡风玻璃","SEA-DGM-006",88.00,2,"OEM New","HUD compatible"),
        ("Doors, Glass & Mirrors","Flush Door Handle L/R","隐藏式门把手","SEA-DGM-007",35.00,4,"OEM New","Electric pop-out"),
        ("Brakes","Front Brake Caliper 6-Pot","前六活塞刹车卡钳","SEA-BRK-001",185.00,1,"OEM New","Performance"),
        ("Brakes","Front Brake Pad Set","前刹车片套装","SEA-BRK-002",22.00,8,"OEM New","High-performance"),
        ("Brakes","Rear Brake Pad Set","后刹车片套装","SEA-BRK-003",18.00,8,"OEM New",""),
        ("Brakes","Front Brake Disc 345mm","前刹车盘345mm","SEA-BRK-004",45.00,4,"OEM New","Drilled vented"),
        ("Brakes","Rear Brake Disc 320mm","后刹车盘320mm","SEA-BRK-005",38.00,4,"OEM New",""),
        ("Suspension & Steering","Front Double Wishbone Kit","前双叉臂悬挂套件","SEA-SUS-001",285.00,1,"OEM New",""),
        ("Suspension & Steering","Rear 5-Link Arm Kit","后五连杆套件","SEA-SUS-002",265.00,1,"OEM New",""),
        ("Suspension & Steering","Front Shock Absorber L","左前减振器","SEA-SUS-003",58.00,2,"OEM New",""),
        ("Suspension & Steering","Rear Shock Absorber L","左后减振器","SEA-SUS-004",52.00,2,"OEM New",""),
        ("Suspension & Steering","EPS Steering Column","EPS转向柱","SEA-SUS-005",145.00,1,"OEM New",""),
        ("EV Drivetrain","OBC 11kW Three-Phase","车载充电机11kW","SEA-EV-001",265.00,1,"OEM New","800V platform"),
        ("EV Drivetrain","Front Motor 230kW","前电机230kW","SEA-EV-002",780.00,1,"OEM New","AWD model"),
        ("EV Drivetrain","Rear Motor 390kW","后电机390kW","SEA-EV-003",950.00,1,"OEM New","Performance model"),
        ("EV Drivetrain","DC-DC 14V Converter","14V DC-DC转换器","SEA-EV-004",125.00,1,"OEM New",""),
        ("EV Drivetrain","BMS Blade Battery CTB","CTB一体化电池BMS","SEA-EV-005",350.00,1,"OEM New","82.56kWh"),
        ("EV Drivetrain","Charging Port 800V CCS","800V充电口","SEA-EV-006",45.00,3,"OEM New","150kW DC max"),
        ("Cooling System","Heat Pump System Assy","热泵系统总成","SEA-COL-001",320.00,1,"OEM New","Integrated"),
        ("Cooling System","Battery Cooling Plate","电池冷却板","SEA-COL-002",145.00,1,"OEM New",""),
        ("Cooling System","Electric Water Pump","电子水泵","SEA-COL-003",32.00,4,"OEM New",""),
        ("Interior","Rotating 15.6in Pad Base","15.6寸旋转中控底座","SEA-INT-001",185.00,2,"OEM New",""),
        ("Interior","Digital Instrument Cluster","数字仪表盘","SEA-INT-002",225.00,1,"OEM New","10.25in"),
        ("Interior","Sporty Floor Mat Set","运动脚垫套装","SEA-INT-003",35.00,6,"OEM New","All-weather"),
        ("Wheels & Tyres","Alloy Wheel 19in","19寸铝合金轮毂","SEA-WHL-001",82.00,4,"OEM New","5×114.3 ET38"),
        ("Wheels & Tyres","Tyre 235/45R19","轮胎 235/45R19","SEA-WHL-002",68.00,4,"OEM New","EV ultra low noise"),
        ("Maintenance","Cabin HEPA Filter","HEPA空调滤芯","SEA-MNT-001",12.00,12,"OEM New",""),
        ("Maintenance","Coolant Concentrate","防冻冷却液原液","SEA-MNT-002",8.00,15,"OEM New","1L OAT type"),
    ]
},

# ═══════════════════════════════════════════════════════════════════════════════
"BYD Dolphin": {
    "info": {"year":"2021–2024","type":"Compact EV Hatchback","voltage":"400V Blade","engine":"EF7A 95kW"},
    "parts": [
        ("Body & Exterior","Front Bumper Assembly","前保险杠总成","DOL-BDY-001",42.00,5,"OEM New",""),
        ("Body & Exterior","Rear Bumper Assembly","后保险杠总成","DOL-BDY-002",38.00,5,"OEM New",""),
        ("Body & Exterior","Hood / Bonnet","发动机盖","DOL-BDY-003",70.00,3,"OEM New",""),
        ("Body & Exterior","Front Left Fender","左前翼子板","DOL-BDY-004",30.00,4,"OEM New",""),
        ("Body & Exterior","Front Right Fender","右前翼子板","DOL-BDY-005",30.00,4,"OEM New",""),
        ("Body & Exterior","Trunk Lid w/ Spoiler","后箱盖带尾翼","DOL-BDY-006",82.00,3,"OEM New",""),
        ("Lighting","Left Headlight LED","左LED大灯","DOL-LGT-001",68.00,2,"OEM New","DRL integrated"),
        ("Lighting","Right Headlight LED","右LED大灯","DOL-LGT-002",68.00,2,"OEM New",""),
        ("Lighting","Left Tail Light","左后尾灯","DOL-LGT-003",45.00,2,"OEM New","C-shape LED"),
        ("Lighting","Right Tail Light","右后尾灯","DOL-LGT-004",45.00,2,"OEM New",""),
        ("Lighting","Rear Fog Lamp","后雾灯","DOL-LGT-005",10.00,6,"OEM New",""),
        ("Doors, Glass & Mirrors","Front Left Door Assembly","左前门总成","DOL-DGM-001",128.00,2,"OEM New",""),
        ("Doors, Glass & Mirrors","Front Right Door Assembly","右前门总成","DOL-DGM-002",128.00,2,"OEM New",""),
        ("Doors, Glass & Mirrors","Rear Left Door Assembly","左后门总成","DOL-DGM-003",118.00,2,"OEM New",""),
        ("Doors, Glass & Mirrors","Rear Right Door Assembly","右后门总成","DOL-DGM-004",118.00,2,"OEM New",""),
        ("Doors, Glass & Mirrors","Windshield (Front)","前挡风玻璃","DOL-DGM-005",58.00,2,"OEM New",""),
        ("Doors, Glass & Mirrors","Electric Left Mirror","左电动后视镜","DOL-DGM-006",35.00,3,"OEM New","Heated"),
        ("Doors, Glass & Mirrors","Electric Right Mirror","右电动后视镜","DOL-DGM-007",35.00,3,"OEM New","Heated"),
        ("Brakes","Front Brake Pad Set","前刹车片套装","DOL-BRK-001",15.00,10,"OEM New",""),
        ("Brakes","Rear Brake Pad Set","后刹车片套装","DOL-BRK-002",13.50,10,"OEM New",""),
        ("Brakes","Front Brake Disc 276mm","前刹车盘276mm","DOL-BRK-003",24.00,6,"OEM New","Vented"),
        ("Brakes","Rear Brake Drum","后制动鼓","DOL-BRK-004",18.00,6,"OEM New","Drum type"),
        ("Suspension & Steering","Front Shock Absorber L","左前减振器","DOL-SUS-001",38.00,2,"OEM New","MacPherson"),
        ("Suspension & Steering","Front Shock Absorber R","右前减振器","DOL-SUS-002",38.00,2,"OEM New",""),
        ("Suspension & Steering","Rear Torsion Beam","后扭力梁","DOL-SUS-003",95.00,1,"OEM New",""),
        ("Suspension & Steering","Front Wheel Bearing L","左前轮毂轴承","DOL-SUS-004",16.00,4,"OEM New",""),
        ("Suspension & Steering","Steering Rack EPS","EPS转向机","DOL-SUS-005",155.00,1,"OEM New",""),
        ("EV Drivetrain","OBC 7kW Charger","7kW车载充电机","DOL-EV-001",195.00,1,"OEM New",""),
        ("EV Drivetrain","BMS Blade Battery","刀片电池BMS","DOL-EV-002",245.00,1,"OEM New","60.48kWh"),
        ("EV Drivetrain","Drive Motor 95kW","95kW驱动电机","DOL-EV-003",520.00,1,"OEM New","Permanent magnet"),
        ("EV Drivetrain","DC-DC Converter","DC-DC转换器","DOL-EV-004",98.00,1,"OEM New","14V"),
        ("EV Drivetrain","Charging Port Type 2","Type2充电口","DOL-EV-005",28.00,5,"OEM New",""),
        ("Cooling System","HVAC Compressor Electric","电动空调压缩机","DOL-COL-001",128.00,2,"OEM New",""),
        ("Cooling System","Battery Thermal Assy","电池热管理总成","DOL-COL-002",168.00,1,"OEM New",""),
        ("Interior","12.8in Central Screen Base","12.8寸中控底座","DOL-INT-001",155.00,1,"OEM New","Rotating"),
        ("Interior","Floor Mat Set","脚垫套装","DOL-INT-002",20.00,10,"OEM New",""),
        ("Wheels & Tyres","Alloy Wheel 17in","17寸铝合金轮毂","DOL-WHL-001",48.00,4,"OEM New","5×100 ET40"),
        ("Wheels & Tyres","Tyre 205/55R17","轮胎 205/55R17","DOL-WHL-002",42.00,4,"OEM New",""),
        ("Maintenance","Cabin Air Filter","空调滤芯","DOL-MNT-001",7.00,20,"OEM New",""),
        ("Maintenance","Wiper Blades","雨刮器","DOL-MNT-002",9.00,10,"OEM New","Pair flat"),
    ]
},

# ═══════════════════════════════════════════════════════════════════════════════
"BYD Han EV": {
    "info": {"year":"2020–2024","type":"Full-size EV Sedan","voltage":"400V Blade","engine":"TZ200XSY AWD 480kW"},
    "parts": [
        ("Body & Exterior","Front Bumper Assembly","前保险杠总成","HAN-BDY-001",72.00,2,"OEM New","Dragon Face design"),
        ("Body & Exterior","Rear Bumper Assembly","后保险杠总成","HAN-BDY-002",68.00,2,"OEM New",""),
        ("Body & Exterior","Hood / Bonnet","发动机盖","HAN-BDY-003",115.00,2,"OEM New","Aluminium alloy"),
        ("Body & Exterior","Front Left Fender","左前翼子板","HAN-BDY-004",52.00,3,"OEM New",""),
        ("Body & Exterior","Front Right Fender","右前翼子板","HAN-BDY-005",52.00,3,"OEM New",""),
        ("Body & Exterior","Trunk Lid w/ Auto","电动行李箱盖","HAN-BDY-006",128.00,1,"OEM New","Auto open/close"),
        ("Body & Exterior","Side Sill Panel L","左门槛饰板","HAN-BDY-007",32.00,3,"OEM New",""),
        ("Lighting","Left LED Matrix Headlight","左LED矩阵大灯","HAN-LGT-001",158.00,1,"OEM New","Adaptive ADB"),
        ("Lighting","Right LED Matrix Headlight","右LED矩阵大灯","HAN-LGT-002",158.00,1,"OEM New",""),
        ("Lighting","Left Tail Light Assembly","左后尾灯总成","HAN-LGT-003",88.00,1,"OEM New","Connected LED bar"),
        ("Lighting","Right Tail Light Assembly","右后尾灯总成","HAN-LGT-004",88.00,1,"OEM New",""),
        ("Lighting","Central LED Light Bar","中央LED灯带","HAN-LGT-005",45.00,2,"OEM New","Front grille"),
        ("Doors, Glass & Mirrors","Front Left Door Panel","左前门板","HAN-DGM-001",175.00,2,"OEM New",""),
        ("Doors, Glass & Mirrors","Front Right Door Panel","右前门板","HAN-DGM-002",175.00,2,"OEM New",""),
        ("Doors, Glass & Mirrors","Rear Left Door Panel","左后门板","HAN-DGM-003",165.00,2,"OEM New",""),
        ("Doors, Glass & Mirrors","Rear Right Door Panel","右后门板","HAN-DGM-004",165.00,2,"OEM New",""),
        ("Doors, Glass & Mirrors","Panoramic Sunroof Glass","全景天窗玻璃","HAN-DGM-005",225.00,1,"OEM New","Sliding"),
        ("Doors, Glass & Mirrors","Front Windshield HUD","HUD前挡玻璃","HAN-DGM-006",95.00,1,"OEM New",""),
        ("Doors, Glass & Mirrors","Power Side Mirror L","左电动后视镜","HAN-DGM-007",52.00,2,"OEM New","Fold+heat+cam"),
        ("Brakes","Front Caliper 4-Pot","前四活塞卡钳","HAN-BRK-001",145.00,1,"OEM New","Brembo supplier"),
        ("Brakes","Front Brake Pad Set","前刹车片套装","HAN-BRK-002",26.00,6,"OEM New","High-perf ceramic"),
        ("Brakes","Rear Brake Pad Set","后刹车片套装","HAN-BRK-003",22.00,6,"OEM New",""),
        ("Brakes","Front Brake Disc 350mm","前刹车盘350mm","HAN-BRK-004",52.00,3,"OEM New","Cross-drilled vented"),
        ("Brakes","EPB Rear Actuator","后EPB执行器","HAN-BRK-005",68.00,2,"OEM New",""),
        ("Suspension & Steering","Front Double Wishbone Set","前双叉臂套件","HAN-SUS-001",320.00,1,"OEM New",""),
        ("Suspension & Steering","Rear Multi-Link Kit","后多连杆套件","HAN-SUS-002",295.00,1,"OEM New",""),
        ("Suspension & Steering","Front Shock Absorber L","左前减振器","HAN-SUS-003",68.00,2,"OEM New","CDC adaptive"),
        ("Suspension & Steering","Rear Shock Absorber L","左后减振器","HAN-SUS-004",62.00,2,"OEM New",""),
        ("Suspension & Steering","EPS Power Steering","EPS助力转向系统","HAN-SUS-005",215.00,1,"OEM New",""),
        ("EV Drivetrain","Front Motor 163kW","前驱电机163kW","HAN-EV-001",720.00,1,"OEM New","AWD model"),
        ("EV Drivetrain","Rear Motor 317kW","后驱电机317kW","HAN-EV-002",880.00,1,"OEM New",""),
        ("EV Drivetrain","OBC 11kW","车载充电机11kW","HAN-EV-003",275.00,1,"OEM New",""),
        ("EV Drivetrain","BMS 100kWh Blade","100kWh刀片电池BMS","HAN-EV-004",380.00,1,"OEM New","85kWh / 100kWh"),
        ("EV Drivetrain","DC-DC 14V High Power","14V大功率DC-DC","HAN-EV-005",138.00,1,"OEM New","3.3kW"),
        ("EV Drivetrain","DC Fast Charge Port","直流快充口","HAN-EV-006",48.00,3,"OEM New","120kW"),
        ("Cooling System","Heat Pump Integrated","一体热泵系统","HAN-COL-001",358.00,1,"OEM New",""),
        ("Cooling System","Battery Cooling Manifold","电池冷却歧管","HAN-COL-002",165.00,1,"OEM New",""),
        ("Interior","15.6in Rotating Central Screen","15.6寸旋转中控屏","HAN-INT-001",285.00,1,"OEM New",""),
        ("Interior","NAPPA Leather Seat FR L","驾驶侧Nappa皮座椅","HAN-INT-002",485.00,1,"OEM New","Ventilation+heat"),
        ("Interior","12-Speaker Dirac Audio Kit","Dirac音响套件","HAN-INT-003",325.00,1,"OEM New","12-speaker"),
        ("Wheels & Tyres","Alloy Wheel 20in","20寸铝合金轮毂","HAN-WHL-001",105.00,4,"OEM New","5×114.3 ET40"),
        ("Wheels & Tyres","Tyre 245/40R20","轮胎 245/40R20","HAN-WHL-002",88.00,4,"OEM New","Low rolling resist."),
        ("Maintenance","HEPA Cabin Filter","HEPA空调滤芯","HAN-MNT-001",14.00,10,"OEM New",""),
        ("Maintenance","Windshield Washer Conc.","玻璃水浓缩液","HAN-MNT-002",4.50,20,"OEM New","500ml"),
    ]
},

# ═══════════════════════════════════════════════════════════════════════════════
"BYD Tang EV": {
    "info": {"year":"2019–2024","type":"Full-size EV SUV 7-seater","voltage":"400V Blade","engine":"AWD 505kW"},
    "parts": [
        ("Body & Exterior","Front Bumper Assembly","前保险杠总成","TNG-BDY-001",85.00,2,"OEM New",""),
        ("Body & Exterior","Rear Bumper Assembly","后保险杠总成","TNG-BDY-002",78.00,2,"OEM New",""),
        ("Body & Exterior","Hood / Bonnet","发动机盖","TNG-BDY-003",128.00,2,"OEM New","Steel"),
        ("Body & Exterior","Front Left Fender","左前翼子板","TNG-BDY-004",58.00,3,"OEM New",""),
        ("Body & Exterior","Front Right Fender","右前翼子板","TNG-BDY-005",58.00,3,"OEM New",""),
        ("Body & Exterior","Tailgate Power","电动尾门","TNG-BDY-006",165.00,1,"OEM New","Hands-free"),
        ("Body & Exterior","Running Board Set","踏板套装","TNG-BDY-007",88.00,2,"OEM New","Auto extending"),
        ("Lighting","Left Adaptive LED Headlight","左自适应LED大灯","TNG-LGT-001",175.00,1,"OEM New",""),
        ("Lighting","Right Adaptive LED Headlight","右自适应LED大灯","TNG-LGT-002",175.00,1,"OEM New",""),
        ("Lighting","Rear LED Light Bar","后LED贯穿灯带","TNG-LGT-003",55.00,2,"OEM New","Full width"),
        ("Lighting","Left Tail Light","左后尾灯","TNG-LGT-004",65.00,2,"OEM New",""),
        ("Lighting","Right Tail Light","右后尾灯","TNG-LGT-005",65.00,2,"OEM New",""),
        ("Doors, Glass & Mirrors","Front Left Door","左前门","TNG-DGM-001",185.00,2,"OEM New",""),
        ("Doors, Glass & Mirrors","Front Right Door","右前门","TNG-DGM-002",185.00,2,"OEM New",""),
        ("Doors, Glass & Mirrors","Rear Left Door","左后门","TNG-DGM-003",175.00,2,"OEM New",""),
        ("Doors, Glass & Mirrors","Rear Right Door","右后门","TNG-DGM-004",175.00,2,"OEM New",""),
        ("Doors, Glass & Mirrors","3rd Row Glass","第三排侧窗玻璃","TNG-DGM-005",48.00,3,"OEM New",""),
        ("Doors, Glass & Mirrors","Panoramic Sunroof Glass","全景天窗玻璃","TNG-DGM-006",245.00,1,"OEM New","2-section"),
        ("Doors, Glass & Mirrors","Side Mirror L w/ Camera","左后视镜带摄像头","TNG-DGM-007",62.00,2,"OEM New","360 surround view"),
        ("Brakes","Front Brake Caliper","前刹车卡钳","TNG-BRK-001",115.00,2,"OEM New",""),
        ("Brakes","Front Brake Pad Set","前刹车片套装","TNG-BRK-002",28.00,6,"OEM New",""),
        ("Brakes","Rear Brake Pad Set","后刹车片套装","TNG-BRK-003",24.00,6,"OEM New",""),
        ("Brakes","Front Brake Disc 356mm","前刹车盘356mm","TNG-BRK-004",58.00,3,"OEM New",""),
        ("Suspension & Steering","Front Air Spring Option","前空气弹簧（可选）","TNG-SUS-001",185.00,1,"OEM New","Adjustable height"),
        ("Suspension & Steering","Front Shock Absorber L","左前减振器","TNG-SUS-002",72.00,2,"OEM New",""),
        ("Suspension & Steering","Rear Shock Absorber L","左后减振器","TNG-SUS-003",65.00,2,"OEM New",""),
        ("Suspension & Steering","Front Subframe","前副车架","TNG-SUS-004",145.00,1,"OEM New",""),
        ("EV Drivetrain","Front Motor 180kW","前驱电机180kW","TNG-EV-001",750.00,1,"OEM New",""),
        ("EV Drivetrain","Rear Motor 325kW","后驱电机325kW","TNG-EV-002",920.00,1,"OEM New",""),
        ("EV Drivetrain","OBC 11kW","车载充电机11kW","TNG-EV-003",275.00,1,"OEM New",""),
        ("EV Drivetrain","BMS 108.8kWh Blade","108.8kWh刀片电池BMS","TNG-EV-004",395.00,1,"OEM New",""),
        ("EV Drivetrain","DC-DC Converter","DC-DC转换器","TNG-EV-005",142.00,1,"OEM New","14V 3.3kW"),
        ("EV Drivetrain","V2L Discharge Module","V2L对外放电模块","TNG-EV-006",88.00,2,"OEM New","3.3kW"),
        ("Cooling System","Heat Pump System","热泵系统总成","TNG-COL-001",375.00,1,"OEM New",""),
        ("Cooling System","Electric Coolant Pump","电子冷却水泵","TNG-COL-002",35.00,4,"OEM New",""),
        ("Interior","2nd Row Captain Seats","第二排商务独立座椅","TNG-INT-001",525.00,1,"OEM New","Pair, massage+vent"),
        ("Interior","3rd Row Bench Seat","第三排长座椅","TNG-INT-002",285.00,1,"OEM New","Fold flat"),
        ("Interior","15.6in Central Screen","15.6寸中控屏","TNG-INT-003",295.00,1,"OEM New","Rotating"),
        ("Interior","Floor Mat 7-seat","7座脚垫套装","TNG-INT-004",42.00,5,"OEM New","3-row"),
        ("Wheels & Tyres","Alloy Wheel 20in","20寸铝合金轮毂","TNG-WHL-001",112.00,4,"OEM New","5×114.3 ET38"),
        ("Wheels & Tyres","Tyre 265/45R20","轮胎 265/45R20","TNG-WHL-002",95.00,4,"OEM New","All-terrain EV"),
        ("Maintenance","Cabin HEPA Filter","HEPA空调滤芯","TNG-MNT-001",14.00,10,"OEM New",""),
        ("Maintenance","Brake Fluid DOT4","刹车油DOT4","TNG-MNT-002",5.50,15,"OEM New",""),
    ]
},

# ═══════════════════════════════════════════════════════════════════════════════
"BYD Song Plus EV": {
    "info": {"year":"2021–2024","type":"Mid-size EV SUV","voltage":"400V Blade","engine":"EF7A 135kW"},
    "parts": [
        ("Body & Exterior","Front Bumper Assembly","前保险杠总成","SPE-BDY-001",55.00,4,"OEM New",""),
        ("Body & Exterior","Rear Bumper Assembly","后保险杠总成","SPE-BDY-002",50.00,4,"OEM New",""),
        ("Body & Exterior","Hood / Bonnet","发动机盖","SPE-BDY-003",88.00,2,"OEM New",""),
        ("Body & Exterior","Front Left Fender","左前翼子板","SPE-BDY-004",38.00,4,"OEM New",""),
        ("Body & Exterior","Front Right Fender","右前翼子板","SPE-BDY-005",38.00,4,"OEM New",""),
        ("Body & Exterior","Power Tailgate","电动尾门","SPE-BDY-006",148.00,1,"OEM New","Auto"),
        ("Lighting","Left LED Headlight","左LED大灯","SPE-LGT-001",88.00,2,"OEM New",""),
        ("Lighting","Right LED Headlight","右LED大灯","SPE-LGT-002",88.00,2,"OEM New",""),
        ("Lighting","Left Tail Light","左后尾灯","SPE-LGT-003",55.00,2,"OEM New",""),
        ("Lighting","Right Tail Light","右后尾灯","SPE-LGT-004",55.00,2,"OEM New",""),
        ("Lighting","Front Fog Lamps Pair","前雾灯一对","SPE-LGT-005",22.00,4,"OEM New",""),
        ("Doors, Glass & Mirrors","Front Left Door","左前门","SPE-DGM-001",145.00,2,"OEM New",""),
        ("Doors, Glass & Mirrors","Front Right Door","右前门","SPE-DGM-002",145.00,2,"OEM New",""),
        ("Doors, Glass & Mirrors","Rear Left Door","左后门","SPE-DGM-003",135.00,2,"OEM New",""),
        ("Doors, Glass & Mirrors","Rear Right Door","右后门","SPE-DGM-004",135.00,2,"OEM New",""),
        ("Doors, Glass & Mirrors","Panoramic Sunroof Glass","全景天窗玻璃","SPE-DGM-005",188.00,1,"OEM New","Fixed large"),
        ("Doors, Glass & Mirrors","Left Side Mirror","左电动后视镜","SPE-DGM-006",42.00,3,"OEM New","Heated fold"),
        ("Doors, Glass & Mirrors","Right Side Mirror","右电动后视镜","SPE-DGM-007",42.00,3,"OEM New",""),
        ("Brakes","Front Brake Pad Set","前刹车片套装","SPE-BRK-001",18.00,8,"OEM New",""),
        ("Brakes","Rear Brake Pad Set","后刹车片套装","SPE-BRK-002",16.00,8,"OEM New",""),
        ("Brakes","Front Brake Disc 316mm","前刹车盘316mm","SPE-BRK-003",32.00,4,"OEM New","Vented"),
        ("Suspension & Steering","Front Shock Absorber L","左前减振器","SPE-SUS-001",48.00,2,"OEM New",""),
        ("Suspension & Steering","Front Shock Absorber R","右前减振器","SPE-SUS-002",48.00,2,"OEM New",""),
        ("Suspension & Steering","Rear Shock Absorber L","左后减振器","SPE-SUS-003",42.00,2,"OEM New",""),
        ("Suspension & Steering","Front Lower Arm L","左前下摆臂","SPE-SUS-004",38.00,3,"OEM New",""),
        ("Suspension & Steering","Steering Rack EPS","EPS转向机","SPE-SUS-005",168.00,1,"OEM New",""),
        ("EV Drivetrain","OBC 7kW","车载充电机7kW","SPE-EV-001",208.00,1,"OEM New",""),
        ("EV Drivetrain","BMS 87.3kWh Blade","87.3kWh刀片电池BMS","SPE-EV-002",315.00,1,"OEM New",""),
        ("EV Drivetrain","Drive Motor 135kW","135kW驱动电机","SPE-EV-003",595.00,1,"OEM New",""),
        ("EV Drivetrain","DC-DC Converter","DC-DC转换器","SPE-EV-004",108.00,1,"OEM New",""),
        ("EV Drivetrain","AC Charging Port Type 2","Type2交流充电口","SPE-EV-005",32.00,4,"OEM New",""),
        ("Cooling System","HVAC EV Compressor","电动空调压缩机","SPE-COL-001",138.00,2,"OEM New",""),
        ("Cooling System","Battery Thermal Assy","电池热管理组件","SPE-COL-002",175.00,1,"OEM New",""),
        ("Interior","12.8in Central Screen","12.8寸中控屏","SPE-INT-001",195.00,1,"OEM New","Rotating"),
        ("Interior","Multi-function Steering Wheel","多功能方向盘","SPE-INT-002",95.00,1,"OEM New","Heated"),
        ("Interior","Floor Mat Set","脚垫套装","SPE-INT-003",26.00,8,"OEM New","5-seat"),
        ("Wheels & Tyres","Alloy Wheel 18in","18寸铝合金轮毂","SPE-WHL-001",65.00,4,"OEM New","5×114.3"),
        ("Wheels & Tyres","Tyre 215/55R18","轮胎 215/55R18","SPE-WHL-002",52.00,4,"OEM New",""),
        ("Maintenance","Cabin Air Filter","空调滤芯","SPE-MNT-001",9.00,12,"OEM New",""),
        ("Maintenance","Wiper Blade Set","雨刮器套装","SPE-MNT-002",10.00,8,"OEM New",""),
    ]
},

# ═══════════════════════════════════════════════════════════════════════════════
"BYD Seagull": {
    "info": {"year":"2023–2024","type":"Mini EV Hatchback","voltage":"400V Blade","engine":"60kW / 75kW"},
    "parts": [
        ("Body & Exterior","Front Bumper Assembly","前保险杠总成","SEG-BDY-001",32.00,6,"OEM New",""),
        ("Body & Exterior","Rear Bumper Assembly","后保险杠总成","SEG-BDY-002",28.00,6,"OEM New",""),
        ("Body & Exterior","Hood / Bonnet","发动机盖","SEG-BDY-003",55.00,4,"OEM New",""),
        ("Body & Exterior","Front Left Fender","左前翼子板","SEG-BDY-004",22.00,5,"OEM New",""),
        ("Body & Exterior","Front Right Fender","右前翼子板","SEG-BDY-005",22.00,5,"OEM New",""),
        ("Body & Exterior","Trunk Lid","行李箱盖","SEG-BDY-006",55.00,3,"OEM New",""),
        ("Lighting","Left Headlight LED","左LED大灯","SEG-LGT-001",45.00,4,"OEM New","Proj. lens DRL"),
        ("Lighting","Right Headlight LED","右LED大灯","SEG-LGT-002",45.00,4,"OEM New",""),
        ("Lighting","Left Tail Light","左后尾灯","SEG-LGT-003",32.00,4,"OEM New",""),
        ("Lighting","Right Tail Light","右后尾灯","SEG-LGT-004",32.00,4,"OEM New",""),
        ("Lighting","Rear Foglight","后雾灯","SEG-LGT-005",8.00,8,"OEM New",""),
        ("Doors, Glass & Mirrors","Front Left Door","左前门","SEG-DGM-001",105.00,3,"OEM New",""),
        ("Doors, Glass & Mirrors","Front Right Door","右前门","SEG-DGM-002",105.00,3,"OEM New",""),
        ("Doors, Glass & Mirrors","Rear Left Door","左后门","SEG-DGM-003",98.00,3,"OEM New",""),
        ("Doors, Glass & Mirrors","Rear Right Door","右后门","SEG-DGM-004",98.00,3,"OEM New",""),
        ("Doors, Glass & Mirrors","Windshield (Front)","前挡风玻璃","SEG-DGM-005",42.00,3,"OEM New",""),
        ("Doors, Glass & Mirrors","Left Side Mirror","左后视镜","SEG-DGM-006",22.00,5,"OEM New","Manual fold"),
        ("Doors, Glass & Mirrors","Right Side Mirror","右后视镜","SEG-DGM-007",22.00,5,"OEM New",""),
        ("Brakes","Front Brake Pad Set","前刹车片套装","SEG-BRK-001",12.00,12,"OEM New",""),
        ("Brakes","Rear Brake Pad Set","后刹车片套装","SEG-BRK-002",11.00,12,"OEM New",""),
        ("Brakes","Front Brake Disc 256mm","前刹车盘256mm","SEG-BRK-003",18.00,6,"OEM New","Vented"),
        ("Brakes","Rear Brake Drum","后制动鼓","SEG-BRK-004",14.00,6,"OEM New",""),
        ("Suspension & Steering","Front Shock Absorber L","左前减振器","SEG-SUS-001",28.00,4,"OEM New","MacPherson"),
        ("Suspension & Steering","Front Shock Absorber R","右前减振器","SEG-SUS-002",28.00,4,"OEM New",""),
        ("Suspension & Steering","Rear Torsion Beam","后扭力梁","SEG-SUS-003",72.00,2,"OEM New",""),
        ("Suspension & Steering","Tie Rod End","横拉杆球头","SEG-SUS-004",8.00,8,"OEM New",""),
        ("Suspension & Steering","Steering Rack EPS","EPS转向机","SEG-SUS-005",128.00,1,"OEM New",""),
        ("EV Drivetrain","OBC 6.6kW","车载充电机6.6kW","SEG-EV-001",165.00,1,"OEM New",""),
        ("EV Drivetrain","Blade Battery BMS 38.88kWh","38.88kWh刀片电池BMS","SEG-EV-002",195.00,1,"OEM New","Std range"),
        ("EV Drivetrain","Drive Motor 60kW","60kW驱动电机","SEG-EV-003",385.00,1,"OEM New",""),
        ("EV Drivetrain","DC-DC Converter","DC-DC转换器","SEG-EV-004",82.00,1,"OEM New","12V"),
        ("EV Drivetrain","AC Charging Port","交流充电口","SEG-EV-005",22.00,6,"OEM New","Type 1 / Type 2"),
        ("Cooling System","HVAC EV Compressor","电动空调压缩机","SEG-COL-001",108.00,2,"OEM New",""),
        ("Cooling System","Battery Thermal Plate","电池冷却板","SEG-COL-002",95.00,2,"OEM New",""),
        ("Interior","10.1in Central Screen","10.1寸中控屏","SEG-INT-001",128.00,1,"OEM New",""),
        ("Interior","Floor Mat Set","脚垫套装","SEG-INT-002",15.00,12,"OEM New","4-pc"),
        ("Interior","Instrument Cluster","仪表盘","SEG-INT-003",68.00,2,"OEM New","5in digital"),
        ("Wheels & Tyres","Alloy Wheel 15in","15寸铝合金轮毂","SEG-WHL-001",32.00,4,"OEM New","4×100 ET40"),
        ("Wheels & Tyres","Tyre 175/60R15","轮胎 175/60R15","SEG-WHL-002",28.00,4,"OEM New","EV spec"),
        ("Maintenance","Cabin Air Filter","空调滤芯","SEG-MNT-001",5.50,20,"OEM New",""),
        ("Maintenance","Wiper Blade Set","雨刮器套装","SEG-MNT-002",7.00,12,"OEM New","Pair"),
    ]
},

} # end MODELS

# ─── Excel builder ────────────────────────────────────────────────────────────

def make_fill(hex_color, fill_type='solid'):
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type=fill_type)

def make_border(style='thin'):
    s = Side(style=style, color='CCCCCC')
    return Border(left=s, right=s, top=s, bottom=s)

def make_font(bold=False, size=10, color='000000', name='Calibri'):
    return Font(bold=bold, size=size, color=color, name=name)

COLS = ['A','B','C','D','E','F','G','H','I']
HEADERS = ['#','Category','Part Name (EN)','Part Name (CN)','Part Number',
           'Unit Price (USD)','MOQ (pcs)','Condition','Notes']
COL_WIDTHS = [5, 22, 32, 22, 18, 16, 11, 14, 28]

def write_model_sheet(wb, model_name, model_data):
    # sanitize sheet name (max 31 chars, no forbidden chars)
    safe = model_name.replace(':','').replace('/','')[:31]
    ws = wb.create_sheet(title=safe)

    # ── Row 1: Model title banner ──────────────────────────────
    ws.merge_cells('A1:I1')
    ws['A1'] = model_name + '  —  BYD OEM Spare Parts  |  CM IMPORT&EXPORT'
    ws['A1'].fill   = make_fill(C_DARK)
    ws['A1'].font   = Font(bold=True, size=14, color=C_WHITE, name='Calibri')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    # ── Row 2: Model info ──────────────────────────────────────
    info = model_data['info']
    info_txt = (f"  Year: {info.get('year','')}   |   Type: {info.get('type','')}"
                f"   |   Voltage: {info.get('voltage','')}   |   Motor: {info.get('engine','')}"
                f"   |   FOB China — All prices in USD")
    ws.merge_cells('A2:I2')
    ws['A2'] = info_txt
    ws['A2'].fill = make_fill(C_ACCENT)
    ws['A2'].font = Font(italic=True, size=9, color='D0E8FF', name='Calibri')
    ws['A2'].alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[2].height = 18

    # ── Row 3: Column headers ──────────────────────────────────
    for ci, (hdr, w) in enumerate(zip(HEADERS, COL_WIDTHS), start=1):
        cell = ws.cell(row=3, column=ci, value=hdr)
        cell.fill = make_fill(C_MID)
        cell.font = Font(bold=True, size=10, color=C_WHITE, name='Calibri')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = make_border()
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[3].height = 22

    # ── Data rows ─────────────────────────────────────────────
    parts = model_data['parts']
    prev_cat = None
    row_num = 4
    idx = 1

    for part in parts:
        cat, name_en, name_cn, part_no, price, moq, cond, notes = part

        # Category separator row
        if cat != prev_cat:
            cat_color = CAT_COLORS.get(cat, "555555")
            ws.merge_cells(f'A{row_num}:I{row_num}')
            cell = ws.cell(row=row_num, column=1, value=f'  {cat}')
            cell.fill = make_fill(cat_color)
            cell.font = Font(bold=True, size=10, color=C_WHITE, name='Calibri')
            cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)
            ws.row_dimensions[row_num].height = 18
            row_num += 1
            prev_cat = cat

        # Alternate row shading
        row_fill = make_fill('FFFFFF') if idx % 2 == 1 else make_fill(C_LIGHT_GRAY)

        values = [idx, cat, name_en, name_cn, part_no, price, moq, cond, notes]
        for ci, val in enumerate(values, start=1):
            cell = ws.cell(row=row_num, column=ci, value=val)
            cell.fill = row_fill
            cell.border = make_border()
            cell.font = make_font(size=9)
            if ci == 1:
                cell.alignment = Alignment(horizontal='center', vertical='center')
            elif ci == 6:  # price
                cell.number_format = '"$"#,##0.00'
                cell.alignment = Alignment(horizontal='right', vertical='center')
                cell.font = Font(bold=True, size=9, color='1A6B2A', name='Calibri')
            elif ci == 7:  # moq
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='left', vertical='center',
                                          wrap_text=(ci in [3,4,9]))
        ws.row_dimensions[row_num].height = 16
        row_num += 1
        idx += 1

    # ── Freeze panes ──────────────────────────────────────────
    ws.freeze_panes = 'A4'

    # ── Auto-filter ───────────────────────────────────────────
    ws.auto_filter.ref = f'A3:I{row_num-1}'

    return ws


def write_index_sheet(wb, models_dict):
    ws = wb.create_sheet(title='INDEX', index=0)

    # Title
    ws.merge_cells('A1:G1')
    ws['A1'] = 'BYD EV Spare Parts Catalog  —  Organized by Vehicle Model'
    ws['A1'].fill = make_fill(C_DARK)
    ws['A1'].font = Font(bold=True, size=15, color=C_WHITE, name='Calibri')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 36

    ws.merge_cells('A2:G2')
    ws['A2'] = 'CM IMPORT&EXPORT  |  FOB China  |  All prices USD  |  Wholesale  |  OEM & Aftermarket'
    ws['A2'].fill = make_fill(C_ACCENT)
    ws['A2'].font = Font(italic=True, size=10, color='D0E8FF', name='Calibri')
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 20

    # Headers
    idx_hdrs = ['#', 'Model', 'Type', 'Year', 'Motor / Power', 'Total Parts', 'Sheet Link']
    idx_widths = [5, 22, 28, 14, 24, 13, 20]
    for ci, (h, w) in enumerate(zip(idx_hdrs, idx_widths), start=1):
        c = ws.cell(row=3, column=ci, value=h)
        c.fill = make_fill(C_BYD_RED)
        c.font = Font(bold=True, size=11, color=C_WHITE, name='Calibri')
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = make_border()
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[3].height = 24

    for i, (mname, mdata) in enumerate(models_dict.items(), start=1):
        row = 3 + i
        info = mdata['info']
        sheet_safe = mname.replace(':','').replace('/','')[:31]
        total_parts = len(mdata['parts'])

        vals = [i, mname, info.get('type',''), info.get('year',''),
                info.get('engine',''), total_parts, f'See sheet: {sheet_safe}']
        row_fill = make_fill('FFFFFF') if i % 2 == 1 else make_fill(C_LIGHT_GRAY)
        for ci, v in enumerate(vals, start=1):
            c = ws.cell(row=row, column=ci, value=v)
            c.fill = row_fill
            c.border = make_border()
            c.font = make_font(size=10)
            c.alignment = Alignment(horizontal='center' if ci in [1,6] else 'left',
                                    vertical='center')
        ws.row_dimensions[row].height = 20

    # Summary row
    total_row = 3 + len(models_dict) + 1
    ws.cell(row=total_row, column=1, value='TOTAL')
    total_parts_sum = sum(len(v['parts']) for v in models_dict.values())
    ws.cell(row=total_row, column=6, value=total_parts_sum)
    for ci in range(1, 8):
        c = ws.cell(row=total_row, column=ci)
        c.fill = make_fill(C_DARK)
        c.font = Font(bold=True, size=11, color=C_WHITE, name='Calibri')
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = make_border()
    ws.row_dimensions[total_row].height = 22

    return ws


def build_workbook():
    wb = openpyxl.Workbook()
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    print('Building INDEX sheet...')
    write_index_sheet(wb, MODELS)

    for model_name, model_data in MODELS.items():
        print(f'  Writing sheet: {model_name} ({len(model_data["parts"])} parts)...')
        write_model_sheet(wb, model_name, model_data)

    wb.save(OUTPUT_FILE)
    total = sum(len(v['parts']) for v in MODELS.values())
    print(f'\n[OK] Saved -> {OUTPUT_FILE}')
    print(f'[OK] {len(MODELS)} model sheets + 1 INDEX = {len(MODELS)+1} sheets total')
    print(f'[OK] {total} parts across all models')


if __name__ == '__main__':
    build_workbook()
