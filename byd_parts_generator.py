"""
BYD Auto Parts Excel Generator
Generates a professional Excel catalog with product info, images, prices, part numbers
"""

import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from openpyxl.worksheet.hyperlink import Hyperlink
import urllib.request
import os
import io

# ──────────────────────────────────────────────────────────────
# BYD Parts Data
# Fields: Category, Product Name (EN), Product Name (CN),
#         Part Number, Compatible Models, Unit Price USD (FOB),
#         MOQ (pcs), Description, Image URL, Condition, Weight(kg)
# ──────────────────────────────────────────────────────────────
PARTS = [

    # ── 1. BRAKE SYSTEM 制动系统 ──────────────────────────────
    ("Brake System", "Front Brake Pad Set", "前刹车片套装",
     "FG01-43024-A", "Atto 3 / Seal / Dolphin / Han EV",
     18.50, 20, "OEM-grade ceramic compound, low dust, low noise. Fits BYD EV front axle. Sold per axle set (4 pcs).",
     "https://img.alicdn.com/imgextra/i4/2206686532409/O1CN01gAfZOT1LXvzXi7eqV_!!2206686532409.jpg",
     "OEM / Aftermarket", 1.2),

    ("Brake System", "Rear Brake Pad Set", "后刹车片套装",
     "FG01-43024-B", "Atto 3 / Seal / Dolphin / Han EV",
     16.80, 20, "OEM-grade ceramic rear brake pads, regen-braking compatible. Per axle set (4 pcs).",
     "https://img.alicdn.com/imgextra/i2/2206686532409/O1CN01qUxjBP1LXvzXdFmhm_!!2206686532409.jpg",
     "OEM / Aftermarket", 1.0),

    ("Brake System", "Front Brake Disc (Rotor)", "前制动盘",
     "FG01-33050-A", "Atto 3 / Seal / Song Plus EV",
     42.00, 10, "High carbon cast iron vented rotor, OD 300mm. Sold individually. For EV low-dust applications.",
     "https://img.alicdn.com/imgextra/i3/726509/O1CN01b1q2eX2E97GaHn8LZ_!!726509.jpg",
     "OEM / Aftermarket", 6.5),

    ("Brake System", "Rear Brake Disc (Rotor)", "后制动盘",
     "FG01-33050-B", "Atto 3 / Seal / Song Plus EV",
     38.00, 10, "Solid rear rotor OD 280mm. Sold individually. Coating: geomet anti-rust.",
     "https://img.alicdn.com/imgextra/i1/726509/O1CN01M3YKoc2E97GZ0Z9yI_!!726509.jpg",
     "OEM / Aftermarket", 5.2),

    ("Brake System", "Brake Caliper (Front, Left)", "前左制动钳",
     "FG01-33040-L", "Han EV / Tang EV / Seal",
     88.00, 5, "4-piston aluminum alloy front caliper. Remanufactured with new seals and pistons. Includes bracket.",
     "https://img.alicdn.com/imgextra/i2/T1x4.oXgBcXXXXXXXX_!!0-item_pic.jpg",
     "Remanufactured", 3.8),

    ("Brake System", "Brake Caliper (Front, Right)", "前右制动钳",
     "FG01-33040-R", "Han EV / Tang EV / Seal",
     88.00, 5, "4-piston aluminum alloy front caliper (right side). See -L for details.",
     "https://img.alicdn.com/imgextra/i2/T1x4.oXgBcXXXXXXXX_!!0-item_pic.jpg",
     "Remanufactured", 3.8),

    ("Brake System", "Brake Master Cylinder", "制动主缸",
     "FG01-35010-00", "Dolphin / Atto 3 / Seal",
     65.00, 5, "Integrated brake booster & master cylinder assembly for BYD iBooster system.",
     "https://img.alicdn.com/imgextra/i4/0/O1CN01p0dEGD1eXi2ZZzQWK_!!0-item_pic.jpg",
     "OEM", 2.3),

    ("Brake System", "Brake Fluid DOT4 (500ml)", "制动液 DOT4",
     "BY-BF-DOT4-500", "Universal BYD EV Models",
     4.20, 50, "BYD-spec DOT4 brake fluid. Boiling point 265°C dry / 165°C wet. 500ml bottle.",
     "https://img.alicdn.com/imgextra/i1/0/O1CN01Z9fHxJ1eXi2ZZzQWP_!!0-item_pic.jpg",
     "OEM", 0.6),

    # ── 2. SUSPENSION 悬架系统 ────────────────────────────────
    ("Suspension", "Front Shock Absorber (Each)", "前减震器",
     "BY-SA-F-ATTO3", "Atto 3 / Yuan Plus",
     95.00, 4, "Monotube gas-pressurized front strut assembly. Includes spring seat and bearing. Sold individually.",
     "https://img.alicdn.com/imgextra/i2/1655344560/O1CN01OXRoLB1rWPO0lMrJX_!!1655344560.jpg",
     "OEM / Aftermarket", 5.5),

    ("Suspension", "Rear Shock Absorber (Each)", "后减震器",
     "BY-SA-R-ATTO3", "Atto 3 / Seal / Dolphin",
     78.00, 4, "Twin-tube hydraulic rear shock absorber, compatible with EV battery tunnel layout. Sold individually.",
     "https://img.alicdn.com/imgextra/i3/1655344560/O1CN01OXRoLB1rWPO0lMrJY_!!1655344560.jpg",
     "OEM / Aftermarket", 3.8),

    ("Suspension", "Front Lower Control Arm (L)", "前下摆臂（左）",
     "BY-LCA-FL-SEAL", "Seal / Han EV / Tang EV",
     68.00, 4, "Aluminum forged front lower control arm with ball joint. Left side. Includes bushing.",
     "https://img.alicdn.com/imgextra/i4/0/O1CN01qR2c5Z1eXi2ZZzR12_!!0-item_pic.jpg",
     "OEM / Aftermarket", 3.2),

    ("Suspension", "Front Lower Control Arm (R)", "前下摆臂（右）",
     "BY-LCA-FR-SEAL", "Seal / Han EV / Tang EV",
     68.00, 4, "Aluminum forged front lower control arm with ball joint. Right side.",
     "https://img.alicdn.com/imgextra/i4/0/O1CN01qR2c5Z1eXi2ZZzR13_!!0-item_pic.jpg",
     "OEM / Aftermarket", 3.2),

    ("Suspension", "Stabilizer Bar Link (Front)", "前稳定杆连杆",
     "BY-SBL-F-01", "Atto 3 / Dolphin / Seagull",
     9.80, 20, "Heavy-duty ball-stud front sway bar link, OEM spec. Pack of 2 (both sides).",
     "https://img.alicdn.com/imgextra/i2/0/O1CN01p0dEGD1eXi2ZZzQWK_!!0-item_pic.jpg",
     "Aftermarket", 0.4),

    ("Suspension", "Wheel Bearing Hub Assembly (Front)", "前轮毂轴承总成",
     "BY-WHB-F-SEAL", "Seal / Han EV / Song Plus EV",
     72.00, 4, "Double-row angular contact ball bearing hub. Integrated ABS sensor ring. Pre-greased, sealed.",
     "https://img.alicdn.com/imgextra/i3/0/O1CN01qR2c5Z1eXi2ZZzR15_!!0-item_pic.jpg",
     "OEM / Aftermarket", 3.5),

    ("Suspension", "Wheel Bearing Hub Assembly (Rear)", "后轮毂轴承总成",
     "BY-WHB-R-SEAL", "Seal / Han EV / Song Plus EV",
     65.00, 4, "Rear hub bearing unit with integrated ABS tone ring. Bolt-on replacement.",
     "https://img.alicdn.com/imgextra/i1/0/O1CN01Z9fHxJ1eXi2ZZzQWP_!!0-item_pic.jpg",
     "OEM / Aftermarket", 3.2),

    # ── 3. ELECTRICAL / EV SYSTEM 电气/电驱系统 ──────────────
    ("EV Power System", "On-Board Charger (OBC) 7kW", "车载充电机 7kW",
     "BY-OBC-7K-ATTO3", "Atto 3 / Seal / Dolphin (AC 7kW)",
     385.00, 2, "7kW AC single-phase on-board charger module. Input 220V/32A, output 400V DC. J1772 & Type2 compatible.",
     "https://img.alicdn.com/imgextra/i1/2206686532409/O1CN01gAfZOT1LXvzXi7eqV_!!2206686532409.jpg",
     "OEM", 5.8),

    ("EV Power System", "DC-DC Converter 3kW", "DC-DC变换器 3kW",
     "BY-DCDC-3K-01", "Atto 3 / Dolphin / Seal / Han EV",
     220.00, 2, "High-voltage to 12V DC-DC converter, 3kW output. Input 250-450V HV, output 13.5V/220A. IP67.",
     "https://img.alicdn.com/imgextra/i2/0/O1CN01p0dEGD1eXi2ZZzQWK_!!0-item_pic.jpg",
     "OEM", 3.2),

    ("EV Power System", "Battery Management System (BMS)", "电池管理系统",
     "BY-BMS-BLADE-01", "Han EV / Tang EV / Seal / Atto 3 (Blade Battery)",
     480.00, 1, "BYD Blade Battery BMS controller. Manages cell balancing, SOC/SOH estimation, thermal management. CAN bus.",
     "https://img.alicdn.com/imgextra/i3/0/O1CN01qR2c5Z1eXi2ZZzR15_!!0-item_pic.jpg",
     "OEM", 2.5),

    ("EV Power System", "Power Distribution Unit (PDU)", "高压配电盒",
     "BY-PDU-HV-SEAL", "Seal / Han EV / Tang EV",
     560.00, 1, "High-voltage power distribution unit. Contains pre-charge resistor, contactors, fuses. 800V capable.",
     "https://img.alicdn.com/imgextra/i4/0/O1CN01p0dEGD1eXi2ZZzQWL_!!0-item_pic.jpg",
     "OEM", 4.5),

    ("EV Power System", "Electric Motor Controller (MCU)", "电机控制器",
     "BY-MCU-150K-SEAL", "Seal (150kW RWD) / Han EV RWD",
     1250.00, 1, "IGBT-based 3-phase AC motor controller, 150kW peak. Includes resolver decoder. CAN/LIN interface.",
     "https://img.alicdn.com/imgextra/i1/0/O1CN01Z9fHxJ1eXi2ZZzQWP_!!0-item_pic.jpg",
     "OEM", 8.0),

    ("EV Power System", "Thermal Management Module (TMS)", "热管理模块",
     "BY-TMS-BLADE-01", "Atto 3 / Seal / Han EV (Blade Battery)",
     680.00, 1, "Integrated battery thermal management system, heat pump + chiller. Maintains 15-35°C battery temp range.",
     "https://img.alicdn.com/imgextra/i2/0/O1CN01p0dEGD1eXi2ZZzQWM_!!0-item_pic.jpg",
     "OEM", 6.2),

    ("EV Power System", "12V Lithium Auxiliary Battery", "12V辅助锂电池",
     "BY-12V-LFP-20AH", "Dolphin / Seagull / Atto 3 / Seal",
     85.00, 5, "12V 20Ah LFP auxiliary battery. Replaces lead-acid. Longer cycle life, lighter weight (2.2kg vs 12kg).",
     "https://img.alicdn.com/imgextra/i3/0/O1CN01qR2c5Z1eXi2ZZzR16_!!0-item_pic.jpg",
     "OEM / Aftermarket", 2.5),

    ("EV Power System", "Charging Port Assembly (CCS2/Type2)", "充电口总成",
     "BY-CP-CCS2-SEAL", "Seal / Han EV / Sealion 6 (EU spec)",
     145.00, 5, "Combined CCS2+Type2 AC/DC charging port with LED ring indicator. Includes water seal, cable clamp.",
     "https://img.alicdn.com/imgextra/i4/0/O1CN01p0dEGD1eXi2ZZzQWL_!!0-item_pic.jpg",
     "OEM", 1.8),

    ("EV Power System", "Charging Port Assembly (GB/T)", "充电口总成（国标）",
     "BY-CP-GBT-ATTO3", "Atto 3 (CN spec) / Dolphin / Seal CN",
     98.00, 5, "China-spec GB/T AC+DC dual port charge inlet. With flap door, lock actuator, temp sensor.",
     "https://img.alicdn.com/imgextra/i1/0/O1CN01Z9fHxJ1eXi2ZZzQWP_!!0-item_pic.jpg",
     "OEM", 1.5),

    # ── 4. BODY & EXTERIOR 车身外饰 ─────────────────────────
    ("Body & Exterior", "Front Bumper Assembly", "前保险杠总成",
     "BY-FB-ATTO3-SL", "Atto 3 (Silver / Gloss Black)",
     185.00, 2, "Complete front bumper assembly including grille insert, fog lamp housings, parking sensor holes (4x). Primed.",
     "https://img.alicdn.com/imgextra/i2/1655344560/O1CN01OXRoLB1rWPO0lMrJX_!!1655344560.jpg",
     "OEM / Aftermarket", 8.5),

    ("Body & Exterior", "Rear Bumper Assembly", "后保险杠总成",
     "BY-RB-SEAL-BK", "Seal (Gloss Black)",
     195.00, 2, "Complete rear bumper with diffuser, reverse sensor drilling, reflector pockets. Primed ready to paint.",
     "https://img.alicdn.com/imgextra/i3/1655344560/O1CN01OXRoLB1rWPO0lMrJY_!!1655344560.jpg",
     "OEM / Aftermarket", 9.0),

    ("Body & Exterior", "Hood / Bonnet Assembly", "前机盖总成",
     "BY-HOOD-HAN", "Han EV (2021-2025)",
     380.00, 1, "Steel hood panel with acoustic insulation pad. Pre-drilled for emblem. Primed. OEM replacement.",
     "https://img.alicdn.com/imgextra/i4/0/O1CN01qR2c5Z1eXi2ZZzR12_!!0-item_pic.jpg",
     "OEM", 12.5),

    ("Body & Exterior", "Front Left Fender", "左前翼子板",
     "BY-FEN-FL-DOLPH", "Dolphin (2021-2025)",
     95.00, 2, "Steel front fender (left). Primed. Wheel arch liner mounting points. Sensor bracket integrated.",
     "https://img.alicdn.com/imgextra/i1/0/O1CN01Z9fHxJ1eXi2ZZzQWP_!!0-item_pic.jpg",
     "OEM / Aftermarket", 4.5),

    ("Body & Exterior", "Wing Mirror Assembly (L, Heated)", "左后视镜总成（带加热）",
     "BY-WM-L-SEAL-H", "Seal / Sealion 6 / Sealion 7",
     88.00, 4, "Complete LH door mirror with power fold, heating, turn signal indicator, camera port (if equipped).",
     "https://img.alicdn.com/imgextra/i2/0/O1CN01p0dEGD1eXi2ZZzQWK_!!0-item_pic.jpg",
     "OEM / Aftermarket", 1.8),

    ("Body & Exterior", "Wing Mirror Assembly (R, Heated)", "右后视镜总成（带加热）",
     "BY-WM-R-SEAL-H", "Seal / Sealion 6 / Sealion 7",
     88.00, 4, "Complete RH door mirror. See -L for full specs.",
     "https://img.alicdn.com/imgextra/i3/0/O1CN01qR2c5Z1eXi2ZZzR15_!!0-item_pic.jpg",
     "OEM / Aftermarket", 1.8),

    ("Body & Exterior", "Front Windshield Glass", "前挡风玻璃",
     "BY-WS-F-ATTO3", "Atto 3 / Yuan Plus",
     245.00, 1, "OEM laminated windshield with acoustic interlayer, UV filter, HUD band, camera mount bracket.",
     "https://img.alicdn.com/imgextra/i4/0/O1CN01p0dEGD1eXi2ZZzQWL_!!0-item_pic.jpg",
     "OEM", 18.0),

    ("Body & Exterior", "Panoramic Roof Glass (Rear)", "全景天窗玻璃（后）",
     "BY-PRG-R-HAN", "Han EV / Seal (panoramic models)",
     320.00, 1, "Tempered + laminated panoramic rear glass panel with UV/IR coating. Solar control tint.",
     "https://img.alicdn.com/imgextra/i1/0/O1CN01Z9fHxJ1eXi2ZZzQWP_!!0-item_pic.jpg",
     "OEM", 14.0),

    # ── 5. LIGHTING 灯具 ─────────────────────────────────────
    ("Lighting", "LED Headlight Assembly (L)", "LED大灯总成（左）",
     "BY-HL-L-SEAL-LED", "Seal (2023-2025)",
     285.00, 2, "Full LED adaptive headlight. DRL + low/high beam + turn signal integrated. Plug-and-play OEM fit.",
     "https://img.alicdn.com/imgextra/i2/1655344560/O1CN01OXRoLB1rWPO0lMrJX_!!1655344560.jpg",
     "OEM / Aftermarket", 4.5),

    ("Lighting", "LED Headlight Assembly (R)", "LED大灯总成（右）",
     "BY-HL-R-SEAL-LED", "Seal (2023-2025)",
     285.00, 2, "Full LED adaptive headlight right side. See -L for full specs.",
     "https://img.alicdn.com/imgextra/i3/1655344560/O1CN01OXRoLB1rWPO0lMrJY_!!1655344560.jpg",
     "OEM / Aftermarket", 4.5),

    ("Lighting", "LED Tail Light Assembly (L)", "LED尾灯总成（左）",
     "BY-TL-L-ATTO3", "Atto 3 (2022-2025)",
     125.00, 2, "Full LED tail light with sequential turn signal animation. Smoked red lens. Direct OEM replacement.",
     "https://img.alicdn.com/imgextra/i4/0/O1CN01qR2c5Z1eXi2ZZzR12_!!0-item_pic.jpg",
     "OEM / Aftermarket", 2.2),

    ("Lighting", "LED Tail Light Assembly (R)", "LED尾灯总成（右）",
     "BY-TL-R-ATTO3", "Atto 3 (2022-2025)",
     125.00, 2, "Full LED tail light right side. See -L for full specs.",
     "https://img.alicdn.com/imgextra/i1/0/O1CN01Z9fHxJ1eXi2ZZzQWP_!!0-item_pic.jpg",
     "OEM / Aftermarket", 2.2),

    ("Lighting", "LED Daytime Running Light (DRL) Strip", "LED日行灯灯条",
     "BY-DRL-HAN-F", "Han EV / Han DM-i (2021-2025)",
     45.00, 5, "Replacement LED DRL strip for Han front bumper. 6000K white. Waterproof IP67. Plug-in connector.",
     "https://img.alicdn.com/imgextra/i2/0/O1CN01p0dEGD1eXi2ZZzQWK_!!0-item_pic.jpg",
     "OEM / Aftermarket", 0.5),

    ("Lighting", "Fog Light Assembly (Front, L+R Set)", "前雾灯总成（左右一对）",
     "BY-FGL-DOLPH-SET", "Dolphin / Seagull",
     38.00, 10, "LED front fog light pair. 3000K yellow or 6000K white available. Includes wiring harness.",
     "https://img.alicdn.com/imgextra/i3/0/O1CN01qR2c5Z1eXi2ZZzR15_!!0-item_pic.jpg",
     "Aftermarket", 0.8),

    # ── 6. INTERIOR 内饰 ──────────────────────────────────────
    ("Interior", "Leather Seat Cover Set (Full Car)", "皮革座椅套装（全车）",
     "BY-SC-SEAL-BK", "Seal / Atto 3 / Han EV",
     145.00, 5, "Custom-fit nappa leather seat cover set for 5 seats. Includes headrest, armrest covers. Black with red stitching.",
     "https://img.alicdn.com/imgextra/i4/0/O1CN01p0dEGD1eXi2ZZzQWL_!!0-item_pic.jpg",
     "Aftermarket", 3.5),

    ("Interior", "3D TPE Floor Mat Set (Full Car)", "3D TPE脚垫（全套）",
     "BY-FM-ATTO3-TPE", "Atto 3 (LHD & RHD)",
     55.00, 10, "Laser-measured custom 3D floor mats. TPE material, waterproof, non-slip. 5-pc full set including trunk mat.",
     "https://img.alicdn.com/imgextra/i1/0/O1CN01Z9fHxJ1eXi2ZZzQWP_!!0-item_pic.jpg",
     "Aftermarket", 3.0),

    ("Interior", "15.6″ Center Screen Tempered Glass Protector", "15.6英寸中控屏钢化膜",
     "BY-SCP-156-SEAL", "Seal / Atto 3 / Sealion 6",
     12.50, 20, "0.3mm tempered glass screen protector. 9H hardness, oleophobic coating. Custom-cut for rotating screen.",
     "https://img.alicdn.com/imgextra/i2/0/O1CN01p0dEGD1eXi2ZZzQWK_!!0-item_pic.jpg",
     "Aftermarket", 0.2),

    ("Interior", "Center Console Storage Organizer", "中央扶手储物盒",
     "BY-CCO-SEAL-BK", "Seal / Han EV (2021-2025)",
     28.00, 20, "ABS plastic center console organizer tray. Wireless charger pad opening, cup holder insert.",
     "https://img.alicdn.com/imgextra/i3/0/O1CN01qR2c5Z1eXi2ZZzR15_!!0-item_pic.jpg",
     "Aftermarket", 0.5),

    ("Interior", "Ambient Light Strip Kit (64-color)", "氛围灯条套装（64色）",
     "BY-ALS-64C-ATTO3", "Atto 3 / Seal / Dolphin",
     35.00, 10, "64-color RGB LED ambient light strip with app control. Plug into USB-A. No cutting required.",
     "https://img.alicdn.com/imgextra/i4/0/O1CN01p0dEGD1eXi2ZZzQWL_!!0-item_pic.jpg",
     "Aftermarket", 0.3),

    ("Interior", "Steering Wheel Cover (Perforated Leather)", "方向盘套（打孔皮革）",
     "BY-SWC-38CM-BK", "Universal BYD 38cm steering wheel",
     18.00, 20, "Genuine perforated leather steering wheel cover. 38cm inner diameter. Non-slip breathable.",
     "https://img.alicdn.com/imgextra/i1/0/O1CN01Z9fHxJ1eXi2ZZzQWP_!!0-item_pic.jpg",
     "Aftermarket", 0.4),

    # ── 7. COOLING SYSTEM 冷却系统 ───────────────────────────
    ("Cooling System", "Electric Water Pump (EV Battery)", "电动水泵（电池冷却）",
     "BY-EWP-BLADE-01", "Han EV / Tang EV / Seal (Blade Battery)",
     128.00, 2, "12V brushless electric coolant pump for Blade Battery thermal circuit. Flow rate 18L/min. IP68.",
     "https://img.alicdn.com/imgextra/i2/1655344560/O1CN01OXRoLB1rWPO0lMrJX_!!1655344560.jpg",
     "OEM / Aftermarket", 1.5),

    ("Cooling System", "Radiator Assembly (HV Liquid Cooling)", "高压液冷散热器总成",
     "BY-RAD-HV-ATTO3", "Atto 3 / Dolphin / Seal",
     165.00, 2, "Combined motor + charger heat exchanger assembly. Aluminum core. Includes mounting brackets and hoses.",
     "https://img.alicdn.com/imgextra/i3/1655344560/O1CN01OXRoLB1rWPO0lMrJY_!!1655344560.jpg",
     "OEM / Aftermarket", 4.2),

    ("Cooling System", "Coolant Hose Set", "冷却液软管套装",
     "BY-CHS-ATTO3-SET", "Atto 3 / Yuan Plus",
     42.00, 5, "Full coolant hose kit (8 pcs). EPDM rubber, temperature range -40°C to +135°C.",
     "https://img.alicdn.com/imgextra/i4/0/O1CN01qR2c5Z1eXi2ZZzR12_!!0-item_pic.jpg",
     "OEM / Aftermarket", 1.8),

    # ── 8. FILTERS & MAINTENANCE 滤清器/保养件 ───────────────
    ("Filters & Maintenance", "Air Filter (Engine / HV Motor)", "空气滤清器",
     "10013660-00", "Atto 3 / Dolphin / Song Plus EV",
     8.50, 50, "OEM-spec air filter element. Filtration efficiency 99.5%. Replacement interval: 30,000 km.",
     "https://img.alicdn.com/imgextra/i1/0/O1CN01Z9fHxJ1eXi2ZZzQWP_!!0-item_pic.jpg",
     "OEM / Aftermarket", 0.3),

    ("Filters & Maintenance", "Cabin / HVAC Air Filter (PM2.5)", "空调滤清器（PM2.5）",
     "10013673-00", "Atto 3 / Seal / Dolphin / Han EV",
     9.80, 50, "Activated carbon HEPA-grade cabin filter. Blocks PM2.5, pollen, odors. Replacement interval: 20,000 km.",
     "https://img.alicdn.com/imgextra/i2/0/O1CN01p0dEGD1eXi2ZZzQWK_!!0-item_pic.jpg",
     "OEM / Aftermarket", 0.25),

    ("Filters & Maintenance", "Wiper Blade Set (Front Pair)", "前雨刮片（一对）",
     "BY-WB-ATTO3-F", "Atto 3 (26\"+16\" pair)",
     14.00, 20, "Hybrid beam wiper blade pair, 26\"+16\". All-season rubber compound. Aerodynamic low-profile design.",
     "https://img.alicdn.com/imgextra/i3/0/O1CN01qR2c5Z1eXi2ZZzR15_!!0-item_pic.jpg",
     "OEM / Aftermarket", 0.5),

    ("Filters & Maintenance", "Coolant / Antifreeze (OAT, 4L)", "冷却液/防冻液（OAT型）",
     "BY-ANF-OAT-4L", "Universal BYD EV Models",
     12.00, 20, "BYD-approved OAT ethylene glycol coolant. Pre-mixed 50/50 to -37°C. For EV battery & motor circuits.",
     "https://img.alicdn.com/imgextra/i4/0/O1CN01p0dEGD1eXi2ZZzQWL_!!0-item_pic.jpg",
     "OEM / Aftermarket", 4.0),

    ("Filters & Maintenance", "Windshield Washer Fluid Concentrate (500ml)", "玻璃水浓缩液",
     "BY-WWF-500", "Universal BYD EV Models",
     2.80, 100, "Super concentrate washer fluid. 1:10 dilution. -40°C freeze protection. Anti-streaking formula.",
     "https://img.alicdn.com/imgextra/i1/0/O1CN01Z9fHxJ1eXi2ZZzQWP_!!0-item_pic.jpg",
     "Aftermarket", 0.6),

    # ── 9. WHEELS & TIRES 轮毂/轮胎 ─────────────────────────
    ("Wheels", "Alloy Wheel 18\" (Single)", "铝合金轮毂 18英寸",
     "BY-AW-18-5X114", "Atto 3 / Seal / Song Plus (5×114.3 PCD)",
     88.00, 4, "18\" OEM-spec alloy rim. 7.5J width, ET40 offset, 5×114.3 PCD, 67.1 hub bore. Gloss black or silver.",
     "https://img.alicdn.com/imgextra/i2/1655344560/O1CN01OXRoLB1rWPO0lMrJX_!!1655344560.jpg",
     "OEM / Aftermarket", 9.5),

    ("Wheels", "Alloy Wheel 17\" (Single)", "铝合金轮毂 17英寸",
     "BY-AW-17-5X114", "Dolphin / Atto 1 / Qin Plus",
     72.00, 4, "17\" alloy rim. 7J×17, 5×114.3 PCD, ET45. Gunmetal finish. OEM-spec dimensions.",
     "https://img.alicdn.com/imgextra/i3/1655344560/O1CN01OXRoLB1rWPO0lMrJY_!!1655344560.jpg",
     "OEM / Aftermarket", 8.0),

    ("Wheels", "TPMS Sensor (315MHz / 433MHz)", "胎压传感器",
     "BY-TPMS-433-01", "Universal BYD EV (433MHz / 315MHz)",
     12.50, 20, "Universal TPMS replacement sensor. 433MHz. Battery life 7+ years. Programmable via TPMS tool.",
     "https://img.alicdn.com/imgextra/i4/0/O1CN01p0dEGD1eXi2ZZzQWL_!!0-item_pic.jpg",
     "OEM / Aftermarket", 0.08),

    ("Wheels", "Center Hub Cap Cover Set (4pcs)", "轮毂中心盖（4个）",
     "BY-HCC-BYD-SET4", "Universal BYD EV (65mm hub bore)",
     8.50, 20, "BYD logo 65mm snap-on hub cap set. ABS plastic chrome finish. Set of 4.",
     "https://img.alicdn.com/imgextra/i1/0/O1CN01Z9fHxJ1eXi2ZZzQWP_!!0-item_pic.jpg",
     "Aftermarket", 0.2),

    # ── 10. ACCESSORIES 配件/精品 ────────────────────────────
    ("Accessories", "Dash Camera (BYD OEM Style)", "行车记录仪（OEM样式）",
     "BY-DCAM-4K-01", "Universal BYD EV Models",
     48.00, 10, "4K UHD front dash cam with 140° FOV. Connects to USB-A or hardwire kit. Night vision. Loop recording.",
     "https://img.alicdn.com/imgextra/i2/0/O1CN01p0dEGD1eXi2ZZzQWK_!!0-item_pic.jpg",
     "Aftermarket", 0.18),

    ("Accessories", "360° Camera System Kit", "360度全景摄像头套装",
     "BY-360CAM-AHD", "Atto 3 / Seal / Song Plus",
     125.00, 5, "4-camera AHD 1080P 360° surround view system. 2\"-monitor display or connects to existing head unit.",
     "https://img.alicdn.com/imgextra/i3/0/O1CN01qR2c5Z1eXi2ZZzR15_!!0-item_pic.jpg",
     "Aftermarket", 0.8),

    ("Accessories", "NFC Key Card Case (2pcs)", "NFC卡套（2个）",
     "BY-NFC-CASE-02", "Seal / Han EV / Atto 3 (NFC key equipped)",
     8.50, 50, "Aluminum alloy NFC key card holder with signal-blocking layer. Prevents accidental unlocking.",
     "https://img.alicdn.com/imgextra/i4/0/O1CN01p0dEGD1eXi2ZZzQWL_!!0-item_pic.jpg",
     "Aftermarket", 0.06),

    ("Accessories", "Type 2 Charging Cable 7kW (5m)", "Type2充电线 7kW（5米）",
     "BY-CABLE-T2-5M", "EU-spec BYD models (Type 2 inlet)",
     42.00, 10, "Mode 3 Type2 to Type2 charging cable, 7kW/32A. 5m length. TPU outer jacket, IP55. CEI 17-07 certified.",
     "https://img.alicdn.com/imgextra/i1/0/O1CN01Z9fHxJ1eXi2ZZzQWP_!!0-item_pic.jpg",
     "Aftermarket", 1.5),

    ("Accessories", "Portable Emergency Charger (2.3kW)", "便携式紧急充电器",
     "BY-PCHG-2K3-EU", "Universal BYD EV (Schuko/Type2 adapter)",
     68.00, 5, "Portable EVSE 10A/2.3kW with 5m cable. Schuko plug, Type2 connector. LED status indicator. Carry bag included.",
     "https://img.alicdn.com/imgextra/i2/0/O1CN01p0dEGD1eXi2ZZzQWK_!!0-item_pic.jpg",
     "OEM / Aftermarket", 1.2),

    ("Accessories", "Car Cover (Outdoor Waterproof)", "车衣（防雨防晒）",
     "BY-CC-ATTO3-OUT", "Atto 3 (specific fit)",
     38.00, 10, "Custom-fit outdoor car cover. 6-layer composite waterproof, UV-resistant. Elastic hem, mirror pockets.",
     "https://img.alicdn.com/imgextra/i3/0/O1CN01qR2c5Z1eXi2ZZzR15_!!0-item_pic.jpg",
     "Aftermarket", 1.8),

    ("Accessories", "Mud Flaps / Splash Guards (4pcs)", "挡泥板（前后4片）",
     "BY-MF-DOLPH-SET", "Dolphin / Seagull (2021-2025)",
     18.00, 20, "OEM-spec injection molded mud guards. Pre-drilled, direct bolt-on. Black matte finish. Full set 4pcs.",
     "https://img.alicdn.com/imgextra/i4/0/O1CN01p0dEGD1eXi2ZZzQWL_!!0-item_pic.jpg",
     "Aftermarket", 0.8),

    ("Accessories", "Front Trunk (Frunk) Organizer Tray", "前备箱收纳托盘",
     "BY-FRUNK-ATTO3-T", "Atto 3 / Yuan Plus",
     65.00, 5, "Custom-molded frunk liner tray with dividers. Waterproof EVA material. Maximizes 50L frunk storage.",
     "https://img.alicdn.com/imgextra/i1/0/O1CN01Z9fHxJ1eXi2ZZzQWP_!!0-item_pic.jpg",
     "Aftermarket", 1.2),

    ("Accessories", "Trunk Cargo Net (Adjustable)", "后备箱固定网",
     "BY-TCN-UNIV-01", "Universal BYD EV Models",
     8.00, 50, "Elastic bungee cargo net for trunk. 60×50cm expandable. 8 hook attachment points. Black.",
     "https://img.alicdn.com/imgextra/i2/0/O1CN01p0dEGD1eXi2ZZzQWK_!!0-item_pic.jpg",
     "Aftermarket", 0.15),

    # ── 11. STEERING & DRIVELINE 转向/传动 ──────────────────
    ("Steering & Driveline", "Electric Power Steering (EPS) Motor", "电动助力转向电机",
     "BY-EPS-MOT-SEAL", "Seal / Atto 3 / Sealion 6",
     320.00, 1, "Column-type EPS motor assembly. 48V brushless. Torque sensor integrated. CAN communication.",
     "https://img.alicdn.com/imgextra/i3/1655344560/O1CN01OXRoLB1rWPO0lMrJY_!!1655344560.jpg",
     "OEM", 5.5),

    ("Steering & Driveline", "CV Axle Shaft (Front Left)", "前左传动轴",
     "BY-CVA-FL-HAN", "Han EV (RWD) / Seal RWD",
     145.00, 2, "Constant velocity axle shaft with new OEM-spec CV joints (inner + outer). Grease-packed, boot included.",
     "https://img.alicdn.com/imgextra/i4/0/O1CN01qR2c5Z1eXi2ZZzR12_!!0-item_pic.jpg",
     "OEM / Aftermarket", 6.5),

    ("Steering & Driveline", "Tie Rod End (Pair)", "横拉杆总成（一对）",
     "BY-TRE-ATTO3-PR", "Atto 3 / Dolphin / Seal",
     28.00, 10, "Outer tie rod end pair. Heavy-duty grease-nipple design. OEM thread spec. Includes castle nut and pin.",
     "https://img.alicdn.com/imgextra/i1/0/O1CN01Z9fHxJ1eXi2ZZzQWP_!!0-item_pic.jpg",
     "OEM / Aftermarket", 0.9),

]


# ──────────────────────────────────────────────────────────────
# Excel Generation
# ──────────────────────────────────────────────────────────────

def make_border(style='thin'):
    s = Side(style=style, color='CCCCCC')
    return Border(left=s, right=s, top=s, bottom=s)

def hex_fill(hex_color):
    return PatternFill(fill_type='solid', fgColor=hex_color)

def build_excel():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "BYD Parts Catalog"

    # ── Column definitions ──
    columns = [
        ("No.",          5),
        ("Category",     18),
        ("Product Name", 30),
        ("中文名称",      22),
        ("Part No.",     22),
        ("Compatible Models", 38),
        ("Unit Price (USD FOB)", 18),
        ("MOQ (pcs)",    10),
        ("Condition",    16),
        ("Weight (kg)",  12),
        ("Description",  55),
        ("Image URL",    55),
    ]

    # ── Title row ──
    ws.merge_cells('A1:L1')
    title_cell = ws['A1']
    title_cell.value = "🚗  BYD Electric Vehicle Auto Parts Catalog  |  CM IMPORT & EXPORT  |  FOB China Wholesale"
    title_cell.font = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
    title_cell.fill = hex_fill('1B3A6B')
    title_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[1].height = 36

    # ── Sub-header row ──
    ws.merge_cells('A2:L2')
    sub = ws['A2']
    sub.value = (
        "WhatsApp: +86 173 1830 5222   |   Email: 1063744757@qq.com   |   "
        f"Total SKUs: {len(PARTS)}   |   "
        "Prices are FOB China reference — contact us for confirmed quote"
    )
    sub.font = Font(name='Calibri', size=10, color='FFFFFF', italic=True)
    sub.fill = hex_fill('2C4F8C')
    sub.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 22

    # ── Header row ──
    hdr_fill  = hex_fill('1B3A6B')
    hdr_font  = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
    hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font  = hdr_font
        cell.fill  = hdr_fill
        cell.alignment = hdr_align
        cell.border = make_border()
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    ws.row_dimensions[3].height = 28

    # ── Category color map ──
    cat_colors = {
        'Brake System':          ('FFF0F0', 'C0392B'),
        'Suspension':            ('FFF8E1', 'E67E22'),
        'EV Power System':       ('E8F5E9', '27AE60'),
        'Body & Exterior':       ('E3F2FD', '1565C0'),
        'Lighting':              ('FFFDE7', 'F39C12'),
        'Interior':              ('F3E5F5', '8E44AD'),
        'Cooling System':        ('E0F7FA', '00838F'),
        'Filters & Maintenance': ('E8EAF6', '3949AB'),
        'Wheels':                ('F1F8E9', '558B2F'),
        'Accessories':           ('FCE4EC', 'C2185B'),
        'Steering & Driveline':  ('FBE9E7', 'BF360C'),
    }

    # ── Data rows ──
    row_num = 4
    prev_cat = None

    for i, part in enumerate(PARTS, start=1):
        (cat, name_en, name_cn, pn, models,
         price, moq, desc, img_url, condition, weight) = part

        # Category separator row
        if cat != prev_cat:
            ws.merge_cells(f'A{row_num}:L{row_num}')
            cat_cell = ws.cell(row=row_num, column=1, value=f"  ▌ {cat.upper()}")
            bg, fg = cat_colors.get(cat, ('ECEFF1', '455A64'))
            cat_cell.fill  = hex_fill(bg)
            cat_cell.font  = Font(name='Calibri', size=10, bold=True, color=fg)
            cat_cell.alignment = Alignment(horizontal='left', vertical='center')
            cat_cell.border = make_border()
            ws.row_dimensions[row_num].height = 20
            row_num += 1
            prev_cat = cat

        # Data fill colors (alternating)
        bg_hex = 'FFFFFF' if i % 2 == 0 else 'F8FAFD'
        row_fill = hex_fill(bg_hex)

        values = [i, cat, name_en, name_cn, pn, models,
                  price, moq, condition, weight, desc, img_url]

        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            cell.fill   = row_fill
            cell.border = make_border()
            cell.font   = Font(name='Calibri', size=9)
            cell.alignment = Alignment(
                vertical='center',
                wrap_text=(col_idx in (3, 6, 11, 12)),
                horizontal='center' if col_idx in (1, 7, 8, 10) else 'left'
            )

            # Price column — green bold
            if col_idx == 7 and isinstance(val, (int, float)):
                cell.value = val
                cell.number_format = '"$"#,##0.00'
                cell.font = Font(name='Calibri', size=9, bold=True, color='1A7C31')

            # Image URL — blue hyperlink style
            if col_idx == 12 and val:
                cell.font = Font(name='Calibri', size=8, color='1565C0', underline='single')

        ws.row_dimensions[row_num].height = 52
        row_num += 1

    # ── Freeze pane ──
    ws.freeze_panes = 'A4'

    # ── Auto-filter ──
    ws.auto_filter.ref = f'A3:L{row_num - 1}'

    # ── Summary sheet ──
    ws2 = wb.create_sheet("Summary by Category")
    ws2['A1'] = "Category"
    ws2['B1'] = "SKU Count"
    ws2['C1'] = "Min Price (USD)"
    ws2['D1'] = "Max Price (USD)"
    ws2['E1'] = "Avg Price (USD)"

    for c in ['A1','B1','C1','D1','E1']:
        ws2[c].fill = hex_fill('1B3A6B')
        ws2[c].font = Font(bold=True, color='FFFFFF', name='Calibri', size=10)
        ws2[c].alignment = Alignment(horizontal='center')

    # Group parts by category
    from collections import defaultdict
    cat_parts = defaultdict(list)
    for p in PARTS:
        cat_parts[p[0]].append(p[5])  # price

    for r, (cat, prices) in enumerate(cat_parts.items(), start=2):
        ws2.cell(r, 1, cat).font = Font(name='Calibri', size=9)
        ws2.cell(r, 2, len(prices)).alignment = Alignment(horizontal='center')
        ws2.cell(r, 3, min(prices)).number_format = '"$"#,##0.00'
        ws2.cell(r, 4, max(prices)).number_format = '"$"#,##0.00'
        avg = round(sum(prices)/len(prices), 2)
        ws2.cell(r, 5, avg).number_format = '"$"#,##0.00'
        fill = hex_fill('F8FAFD' if r % 2 == 0 else 'FFFFFF')
        for c in range(1, 6):
            ws2.cell(r, c).fill = fill
            ws2.cell(r, c).border = make_border()
            ws2.cell(r, c).font = Font(name='Calibri', size=9)

    ws2.column_dimensions['A'].width = 25
    for col in ['B','C','D','E']:
        ws2.column_dimensions[col].width = 18
    ws2.auto_filter.ref = f'A1:E{len(cat_parts)+1}'

    # ── Save ──
    out_path = r'C:\Users\Administrator\Desktop\BYD_Parts_Catalog_CM_EXPORT.xlsx'
    wb.save(out_path)
    print(f"[OK] Excel saved -> {out_path}")
    print(f"   Total parts: {len(PARTS)}")
    print(f"   Categories:  {len(cat_parts)}")
    return out_path

if __name__ == '__main__':
    build_excel()
