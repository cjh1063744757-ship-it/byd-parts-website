"""
Sinotruk (中国重汽) Spare Parts Catalog — Organized by Vehicle Model
CM IMPORT&EXPORT
Output: Sinotruk_Parts_By_Model_CM_EXPORT.xlsx
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT_FILE = r'C:\Users\Administrator\Desktop\Sinotruk_Parts_By_Model_CM_EXPORT.xlsx'

# ─── Colour palette ──────────────────────────────────────────────────────────
C_RED_DARK   = "8B0000"   # 重汽深红
C_RED_MID    = "B71C1C"
C_DARK       = "1A1A2E"
C_ACCENT     = "0D3B6E"
C_WHITE      = "FFFFFF"
C_LIGHT_GRAY = "F4F4F4"

# System category colours
CAT_COLORS = {
    "Cabin & Body"             : "1565C0",
    "Engine & Fuel System"     : "BF360C",
    "Transmission & Drivetrain": "4A148C",
    "Axle & Differential"      : "1B5E20",
    "Brakes & Air System"      : "B71C1C",
    "Suspension & Steering"    : "0E6655",
    "Electrical & Lighting"    : "F57F17",
    "Exhaust & Emission"       : "4E342E",
    "Cooling System"           : "01579B",
    "Hydraulics"               : "33691E",
    "Wheels & Tyres"           : "E65100",
    "Filters & Maintenance"    : "37474F",
}

# ─── Model Data ──────────────────────────────────────────────────────────────
# Tuple: (category, name_en, name_cn, part_no, price_usd, moq, condition, notes)

MODELS = {

# ═══════════════════════════════════════════════════════════════════════════════
"HOWO 6×4 Tractor (371HP)": {
    "info": {
        "year": "2010–2024", "type": "Heavy Tractor Truck 6×4",
        "engine": "WD615/WD10 371HP", "gvw": "GVW 25T"
    },
    "parts": [
        # Cabin & Body
        ("Cabin & Body","Left Door Assembly","左车门总成","HW371-CAB-001",128.00,2,"OEM New","Complete with glass"),
        ("Cabin & Body","Right Door Assembly","右车门总成","HW371-CAB-002",128.00,2,"OEM New",""),
        ("Cabin & Body","Front Bumper Steel","前保险杠钢制","HW371-CAB-003",72.00,3,"OEM New","Heavy duty"),
        ("Cabin & Body","Windshield Glass","前挡风玻璃","HW371-CAB-004",68.00,2,"OEM New","Laminated"),
        ("Cabin & Body","Cabin Roof Panel","驾驶室顶盖","HW371-CAB-005",115.00,1,"OEM New","High roof option"),
        ("Cabin & Body","Left Side Mirror Assy","左后视镜总成","HW371-CAB-006",32.00,4,"OEM New","Heated"),
        ("Cabin & Body","Right Side Mirror Assy","右后视镜总成","HW371-CAB-007",32.00,4,"OEM New","Heated"),
        ("Cabin & Body","Engine Hood","发动机盖","HW371-CAB-008",95.00,2,"OEM New","Steel"),
        ("Cabin & Body","Rear-View Camera Kit","倒车摄像头套件","HW371-CAB-009",28.00,5,"Aftermarket",""),
        ("Cabin & Body","Cabin Mounting Bracket","驾驶室支架总成","HW371-CAB-010",55.00,3,"OEM New","Front air bag type"),
        # Engine & Fuel System
        ("Engine & Fuel System","Fuel Injector WD615","WD615喷油嘴","HW371-ENG-001",38.00,6,"OEM New","Bosch spec"),
        ("Engine & Fuel System","High-Pressure Fuel Pump","高压油泵总成","HW371-ENG-002",195.00,1,"OEM New","VE pump"),
        ("Engine & Fuel System","Turbocharger HX40W","HX40W涡轮增压器","HW371-ENG-003",185.00,2,"OEM New","Holset brand"),
        ("Engine & Fuel System","Engine Oil Pump","机油泵","HW371-ENG-004",45.00,3,"OEM New",""),
        ("Engine & Fuel System","Piston & Ring Set","活塞及活塞环套件","HW371-ENG-005",62.00,6,"OEM New","STD size"),
        ("Engine & Fuel System","Cylinder Head Gasket Set","缸盖衬垫套件","HW371-ENG-006",28.00,6,"OEM New","Full set"),
        ("Engine & Fuel System","Crankshaft Main Bearing","曲轴主轴承","HW371-ENG-007",18.00,10,"OEM New","STD size"),
        ("Engine & Fuel System","Fuel Feed Pump","输油泵","HW371-ENG-008",32.00,4,"OEM New",""),
        ("Engine & Fuel System","Engine Valve Set","气门套件","HW371-ENG-009",22.00,10,"OEM New","Inlet+Exhaust"),
        # Transmission & Drivetrain
        ("Transmission & Drivetrain","Gearbox HW19710","HW19710变速箱总成","HW371-TRN-001",1250.00,1,"OEM New","10-speed"),
        ("Transmission & Drivetrain","Clutch Disc 430mm","430mm离合器片","HW371-TRN-002",48.00,5,"OEM New",""),
        ("Transmission & Drivetrain","Clutch Pressure Plate","离合器压盘","HW371-TRN-003",72.00,3,"OEM New",""),
        ("Transmission & Drivetrain","Clutch Release Bearing","离合器分离轴承","HW371-TRN-004",18.00,8,"OEM New",""),
        ("Transmission & Drivetrain","Drive Shaft Front Section","传动轴前节","HW371-TRN-005",145.00,2,"OEM New",""),
        ("Transmission & Drivetrain","Drive Shaft Rear Section","传动轴后节","HW371-TRN-006",138.00,2,"OEM New",""),
        ("Transmission & Drivetrain","PTO Power Take-Off","取力器总成","HW371-TRN-007",195.00,1,"OEM New",""),
        # Axle & Differential
        ("Axle & Differential","Front Axle Assembly HC16","HC16前桥总成","HW371-AXL-001",680.00,1,"OEM New","16T rating"),
        ("Axle & Differential","Rear Drive Axle HC16","HC16后驱桥总成","HW371-AXL-002",950.00,1,"OEM New",""),
        ("Axle & Differential","Differential Assembly","差速器总成","HW371-AXL-003",285.00,1,"OEM New",""),
        ("Axle & Differential","Wheel Hub Front","前轮毂","HW371-AXL-004",65.00,4,"OEM New",""),
        ("Axle & Differential","Wheel Hub Rear","后轮毂","HW371-AXL-005",72.00,4,"OEM New",""),
        ("Axle & Differential","King Pin Set","主销套件","HW371-AXL-006",18.00,10,"OEM New","With bushing"),
        # Brakes & Air System
        ("Brakes & Air System","Front Brake Chamber","前制动气室","HW371-BRK-001",28.00,8,"OEM New","T30 type"),
        ("Brakes & Air System","Rear Spring Brake Chamber","后弹簧制动气室","HW371-BRK-002",45.00,6,"OEM New","T30/30"),
        ("Brakes & Air System","Brake Drum Front","前制动鼓","HW371-BRK-003",52.00,4,"OEM New",""),
        ("Brakes & Air System","Brake Drum Rear","后制动鼓","HW371-BRK-004",55.00,4,"OEM New",""),
        ("Brakes & Air System","Brake Lining Shoe Set","刹车蹄片套件","HW371-BRK-005",22.00,8,"OEM New","4pcs"),
        ("Brakes & Air System","Air Compressor","空气压缩机","HW371-BRK-006",145.00,2,"OEM New",""),
        ("Brakes & Air System","ABS Sensor","ABS传感器","HW371-BRK-007",18.00,8,"OEM New",""),
        ("Brakes & Air System","Slack Adjuster Automatic","自动调整臂","HW371-BRK-008",22.00,8,"OEM New",""),
        # Suspension & Steering
        ("Suspension & Steering","Front Leaf Spring Set","前板簧总成","HW371-SUS-001",95.00,2,"OEM New","8-leaf"),
        ("Suspension & Steering","Rear Leaf Spring Set","后板簧总成","HW371-SUS-002",115.00,2,"OEM New","12-leaf"),
        ("Suspension & Steering","Front Shock Absorber","前减振器","HW371-SUS-003",35.00,4,"OEM New",""),
        ("Suspension & Steering","Steering Gear Box","转向机总成","HW371-SUS-004",285.00,1,"OEM New","ZF type"),
        ("Suspension & Steering","Tie Rod Assembly","横拉杆总成","HW371-SUS-005",48.00,4,"OEM New",""),
        ("Suspension & Steering","Drag Link","直拉杆","HW371-SUS-006",42.00,4,"OEM New",""),
        ("Suspension & Steering","U-Bolt Set Front","前U形螺栓套件","HW371-SUS-007",12.00,20,"OEM New",""),
        # Electrical & Lighting
        ("Electrical & Lighting","Left Headlight Assy","左大灯总成","HW371-ELC-001",42.00,4,"OEM New","H7 bulb"),
        ("Electrical & Lighting","Right Headlight Assy","右大灯总成","HW371-ELC-002",42.00,4,"OEM New",""),
        ("Electrical & Lighting","Alternator 28V 55A","28V 55A发电机","HW371-ELC-003",125.00,2,"OEM New",""),
        ("Electrical & Lighting","Starter Motor 24V","24V起动机","HW371-ELC-004",115.00,2,"OEM New","6kW"),
        ("Electrical & Lighting","Battery 24V 180Ah","24V 180Ah蓄电池","HW371-ELC-005",85.00,2,"OEM New","Pair"),
        ("Electrical & Lighting","Dashboard Instrument","仪表总成","HW371-ELC-006",95.00,2,"OEM New",""),
        ("Electrical & Lighting","Wiper Motor Assembly","雨刮电机总成","HW371-ELC-007",38.00,4,"OEM New",""),
        # Exhaust & Emission
        ("Exhaust & Emission","Exhaust Pipe Assembly","排气管总成","HW371-EXH-001",65.00,2,"OEM New",""),
        ("Exhaust & Emission","Muffler / Silencer","消声器总成","HW371-EXH-002",55.00,3,"OEM New",""),
        ("Exhaust & Emission","DPF Diesel Particle Filter","柴油颗粒捕捉器","HW371-EXH-003",285.00,1,"OEM New","Euro IV/V"),
        ("Exhaust & Emission","EGR Valve","EGR废气再循环阀","HW371-EXH-004",78.00,3,"OEM New",""),
        # Cooling System
        ("Cooling System","Radiator Assembly","散热器总成","HW371-COL-001",145.00,2,"OEM New","Aluminium core"),
        ("Cooling System","Fan Clutch","风扇离合器","HW371-COL-002",85.00,3,"OEM New","Silicon oil type"),
        ("Cooling System","Water Pump","水泵","HW371-COL-003",45.00,4,"OEM New",""),
        ("Cooling System","Thermostat 82°C","节温器82°C","HW371-COL-004",12.00,15,"OEM New",""),
        ("Cooling System","Intercooler","中冷器","HW371-COL-005",95.00,2,"OEM New",""),
        # Wheels & Tyres
        ("Wheels & Tyres","Steel Wheel Rim 22.5in","22.5寸钢制轮辋","HW371-WHL-001",48.00,8,"OEM New","10×22.5"),
        ("Wheels & Tyres","Tyre 12R22.5","轮胎 12R22.5","HW371-WHL-002",168.00,6,"OEM New","Drive axle spec"),
        ("Wheels & Tyres","Tyre 315/80R22.5","轮胎 315/80R22.5","HW371-WHL-003",185.00,6,"OEM New","Steer axle"),
        ("Wheels & Tyres","Wheel Nut M22","M22轮毂螺母","HW371-WHL-004",1.80,100,"OEM New","Set of 10"),
        # Filters & Maintenance
        ("Filters & Maintenance","Engine Oil Filter","机油滤清器","HW371-MNT-001",8.50,20,"OEM New",""),
        ("Filters & Maintenance","Fuel Filter Primary","初级燃油滤清器","HW371-MNT-002",6.00,20,"OEM New",""),
        ("Filters & Maintenance","Fuel Filter Secondary","精级燃油滤清器","HW371-MNT-003",8.00,20,"OEM New",""),
        ("Filters & Maintenance","Air Filter Primary","初级空气滤清器","HW371-MNT-004",12.00,10,"OEM New",""),
        ("Filters & Maintenance","Air Filter Secondary","安全芯","HW371-MNT-005",8.00,10,"OEM New",""),
        ("Filters & Maintenance","Cabin Air Filter","驾驶室空调滤芯","HW371-MNT-006",6.00,20,"OEM New",""),
        ("Filters & Maintenance","Engine Oil 15W-40 7L","15W-40机油7升","HW371-MNT-007",18.00,10,"OEM New","API CI-4"),
        ("Filters & Maintenance","Coolant Antifreeze","防冻冷却液","HW371-MNT-008",9.00,12,"OEM New","2L OAT"),
    ]
},

# ═══════════════════════════════════════════════════════════════════════════════
"HOWO 6×4 Dump Truck": {
    "info": {
        "year": "2010–2024", "type": "Heavy Dump Truck 6×4",
        "engine": "WD615 336HP / WD10 380HP", "gvw": "GVW 31T"
    },
    "parts": [
        # Cabin & Body
        ("Cabin & Body","Left Door Assembly","左车门总成","HWDP-CAB-001",125.00,2,"OEM New",""),
        ("Cabin & Body","Right Door Assembly","右车门总成","HWDP-CAB-002",125.00,2,"OEM New",""),
        ("Cabin & Body","Front Bumper Heavy","前重型保险杠","HWDP-CAB-003",88.00,2,"OEM New","With steps"),
        ("Cabin & Body","Windshield Glass","前挡风玻璃","HWDP-CAB-004",68.00,2,"OEM New","Laminated"),
        ("Cabin & Body","Left Side Mirror w/ Heat","左电热后视镜","HWDP-CAB-005",35.00,4,"OEM New",""),
        ("Cabin & Body","Right Side Mirror w/ Heat","右电热后视镜","HWDP-CAB-006",35.00,4,"OEM New",""),
        ("Cabin & Body","Engine Hood Assembly","发动机盖总成","HWDP-CAB-007",95.00,2,"OEM New",""),
        ("Cabin & Body","Dump Body 20CBM Steel","20方钢制货箱","HWDP-CAB-008",2850.00,1,"OEM New","6mm floor 4mm wall"),
        ("Cabin & Body","Tailgate Dump Body","货箱后挡板","HWDP-CAB-009",185.00,1,"OEM New","With lock"),
        # Engine & Fuel System
        ("Engine & Fuel System","Fuel Injector WD615","WD615喷油嘴","HWDP-ENG-001",38.00,6,"OEM New","Bosch spec"),
        ("Engine & Fuel System","Turbocharger HX35","HX35涡轮增压器","HWDP-ENG-002",165.00,2,"OEM New","Holset"),
        ("Engine & Fuel System","High-Pressure Fuel Pump","高压油泵","HWDP-ENG-003",195.00,1,"OEM New","VE pump"),
        ("Engine & Fuel System","Engine Oil Pump","机油泵","HWDP-ENG-004",45.00,3,"OEM New",""),
        ("Engine & Fuel System","Piston & Ring Set","活塞及活塞环","HWDP-ENG-005",62.00,6,"OEM New","STD size"),
        ("Engine & Fuel System","Cylinder Head Gasket Set","缸盖垫套件","HWDP-ENG-006",28.00,6,"OEM New",""),
        ("Engine & Fuel System","Connecting Rod Bearing","连杆轴承","HWDP-ENG-007",12.00,12,"OEM New","STD"),
        ("Engine & Fuel System","Fuel Tank 300L","300升燃油箱","HWDP-ENG-008",185.00,1,"OEM New","Aluminium alloy"),
        # Transmission & Drivetrain
        ("Transmission & Drivetrain","Gearbox HW19712","HW19712变速箱总成","HWDP-TRN-001",1350.00,1,"OEM New","12-speed overdrive"),
        ("Transmission & Drivetrain","Clutch Disc 430mm","430mm离合器片","HWDP-TRN-002",48.00,5,"OEM New",""),
        ("Transmission & Drivetrain","Clutch Pressure Plate","离合器压盘","HWDP-TRN-003",72.00,3,"OEM New",""),
        ("Transmission & Drivetrain","Drive Shaft Center Section","传动轴中节","HWDP-TRN-004",165.00,2,"OEM New",""),
        ("Transmission & Drivetrain","Inter-Axle Differential","中桥差速器","HWDP-TRN-005",265.00,1,"OEM New",""),
        # Axle & Differential
        ("Axle & Differential","Front Axle HC16 Assy","HC16前桥总成","HWDP-AXL-001",680.00,1,"OEM New","16T"),
        ("Axle & Differential","Mid Rear Axle HC13 Assy","HC13中后桥总成","HWDP-AXL-002",880.00,1,"OEM New","13T each"),
        ("Axle & Differential","Rear-Rear Axle HC13 Assy","HC13后后桥总成","HWDP-AXL-003",880.00,1,"OEM New",""),
        ("Axle & Differential","Wheel Hub Front","前轮毂","HWDP-AXL-004",65.00,4,"OEM New",""),
        ("Axle & Differential","Wheel Hub Rear","后轮毂","HWDP-AXL-005",72.00,4,"OEM New",""),
        ("Axle & Differential","Front King Pin Set","前桥主销套件","HWDP-AXL-006",18.00,10,"OEM New",""),
        # Brakes & Air System
        ("Brakes & Air System","Front Brake Chamber T30","前制动气室T30","HWDP-BRK-001",28.00,8,"OEM New",""),
        ("Brakes & Air System","Rear Spring Brake Chamber","后弹簧制动气室","HWDP-BRK-002",45.00,6,"OEM New","T30/30"),
        ("Brakes & Air System","Brake Drum Front 420mm","前制动鼓420mm","HWDP-BRK-003",52.00,4,"OEM New",""),
        ("Brakes & Air System","Brake Drum Rear 420mm","后制动鼓420mm","HWDP-BRK-004",55.00,4,"OEM New",""),
        ("Brakes & Air System","Brake Lining Set","制动蹄片套件","HWDP-BRK-005",22.00,8,"OEM New",""),
        ("Brakes & Air System","Air Compressor","空气压缩机","HWDP-BRK-006",145.00,2,"OEM New",""),
        ("Brakes & Air System","Air Dryer","空气干燥器","HWDP-BRK-007",85.00,3,"OEM New",""),
        ("Brakes & Air System","Slack Adjuster Auto","自动调整臂","HWDP-BRK-008",22.00,8,"OEM New",""),
        # Suspension & Steering
        ("Suspension & Steering","Front Leaf Spring Set","前板簧","HWDP-SUS-001",98.00,2,"OEM New",""),
        ("Suspension & Steering","Rear Leaf Spring Bogie","后平衡轴板簧","HWDP-SUS-002",145.00,2,"OEM New","Bogie type"),
        ("Suspension & Steering","Balance Shaft Assembly","平衡轴总成","HWDP-SUS-003",285.00,1,"OEM New",""),
        ("Suspension & Steering","Steering Gear Box ZF","ZF转向机总成","HWDP-SUS-004",285.00,1,"OEM New",""),
        ("Suspension & Steering","Drag Link","直拉杆","HWDP-SUS-005",42.00,4,"OEM New",""),
        # Electrical & Lighting
        ("Electrical & Lighting","Left Headlight Assy","左大灯总成","HWDP-ELC-001",42.00,4,"OEM New",""),
        ("Electrical & Lighting","Right Headlight Assy","右大灯总成","HWDP-ELC-002",42.00,4,"OEM New",""),
        ("Electrical & Lighting","Alternator 28V 55A","发电机28V 55A","HWDP-ELC-003",125.00,2,"OEM New",""),
        ("Electrical & Lighting","Starter Motor 24V 6kW","起动机24V 6kW","HWDP-ELC-004",115.00,2,"OEM New",""),
        ("Electrical & Lighting","Dashboard Cluster","仪表板总成","HWDP-ELC-005",95.00,2,"OEM New",""),
        # Hydraulics (dump specific)
        ("Hydraulics","Hydraulic Pump","液压泵","HWDP-HYD-001",185.00,2,"OEM New","CBT-F425 gear pump"),
        ("Hydraulics","Hydraulic Cylinder 3-Stage","三级举升油缸","HWDP-HYD-002",385.00,1,"OEM New","380mm stroke"),
        ("Hydraulics","Hydraulic Control Valve","液压控制阀","HWDP-HYD-003",68.00,3,"OEM New",""),
        ("Hydraulics","Hydraulic Oil Tank","液压油箱","HWDP-HYD-004",75.00,2,"OEM New","80L"),
        ("Hydraulics","Hydraulic Hose Kit","液压胶管套件","HWDP-HYD-005",35.00,5,"OEM New","High pressure"),
        ("Hydraulics","Hydraulic Oil Filter","液压油滤清器","HWDP-HYD-006",12.00,15,"OEM New",""),
        ("Hydraulics","PTO Gear Box","取力器变速箱","HWDP-HYD-007",195.00,1,"OEM New",""),
        # Cooling
        ("Cooling System","Radiator Assembly","散热器总成","HWDP-COL-001",145.00,2,"OEM New","Al core"),
        ("Cooling System","Fan Clutch Silicone","硅油风扇离合器","HWDP-COL-002",85.00,3,"OEM New",""),
        ("Cooling System","Water Pump","水泵","HWDP-COL-003",45.00,4,"OEM New",""),
        # Wheels & Tyres
        ("Wheels & Tyres","Steel Wheel Rim 22.5in","22.5寸钢圈","HWDP-WHL-001",48.00,10,"OEM New","10H×22.5"),
        ("Wheels & Tyres","Tyre 12.00R20 Drive","12.00R20驱动轮胎","HWDP-WHL-002",155.00,6,"OEM New",""),
        ("Wheels & Tyres","Tyre 12R22.5 Steer","12R22.5转向轮胎","HWDP-WHL-003",168.00,4,"OEM New",""),
        # Filters & Maintenance
        ("Filters & Maintenance","Engine Oil Filter","机油滤清器","HWDP-MNT-001",8.50,20,"OEM New",""),
        ("Filters & Maintenance","Fuel Filter Primary","初级油水分离器","HWDP-MNT-002",6.00,20,"OEM New",""),
        ("Filters & Maintenance","Air Filter Primary","空气滤清器","HWDP-MNT-003",12.00,10,"OEM New",""),
        ("Filters & Maintenance","Hydraulic Oil Filter","液压油滤芯","HWDP-MNT-004",10.00,15,"OEM New",""),
        ("Filters & Maintenance","Gearbox Oil 90W","变速箱油90W","HWDP-MNT-005",22.00,8,"OEM New","5L GL-5"),
    ]
},

# ═══════════════════════════════════════════════════════════════════════════════
"HOWO A7 6×4": {
    "info": {
        "year": "2012–2024", "type": "Premium Tractor / Dump 6×4",
        "engine": "MC11/WD10 380–420HP", "gvw": "GVW 25–31T"
    },
    "parts": [
        # Cabin & Body
        ("Cabin & Body","Left Door A7 Style","A7左车门总成","HWA7-CAB-001",148.00,2,"OEM New","A7 series"),
        ("Cabin & Body","Right Door A7 Style","A7右车门总成","HWA7-CAB-002",148.00,2,"OEM New",""),
        ("Cabin & Body","Front Bumper Chrome","前镀铬保险杠","HWA7-CAB-003",98.00,2,"OEM New","With LED strip"),
        ("Cabin & Body","Windshield A7","A7前挡风玻璃","HWA7-CAB-004",78.00,2,"OEM New","Heated option"),
        ("Cabin & Body","Roof Sun Visor","顶置遮阳板","HWA7-CAB-005",45.00,4,"OEM New",""),
        ("Cabin & Body","Left Side Mirror A7","A7左后视镜","HWA7-CAB-006",42.00,3,"OEM New","Electric fold+heat"),
        ("Cabin & Body","Right Side Mirror A7","A7右后视镜","HWA7-CAB-007",42.00,3,"OEM New",""),
        ("Cabin & Body","Engine Hood A7","A7发动机盖","HWA7-CAB-008",105.00,2,"OEM New",""),
        ("Cabin & Body","Sleeping Berth Upper","上铺总成","HWA7-CAB-009",185.00,1,"OEM New","Fold-down"),
        ("Cabin & Body","Door Seal Strip Set","门密封条套件","HWA7-CAB-010",22.00,6,"OEM New","4pcs"),
        # Engine & Fuel System
        ("Engine & Fuel System","MC11 Common Rail Injector","MC11共轨喷射器","HWA7-ENG-001",68.00,6,"OEM New","Bosch CRIN"),
        ("Engine & Fuel System","High-Pressure Common Rail Pump","高压共轨泵","HWA7-ENG-002",285.00,1,"OEM New","Bosch CP3"),
        ("Engine & Fuel System","Turbocharger HX55","HX55涡轮增压器","HWA7-ENG-003",225.00,1,"OEM New","Variable geometry"),
        ("Engine & Fuel System","EGR Cooler","EGR冷却器","HWA7-ENG-004",85.00,2,"OEM New",""),
        ("Engine & Fuel System","Piston Kit MC11","MC11活塞总成","HWA7-ENG-005",75.00,6,"OEM New",""),
        ("Engine & Fuel System","Cylinder Head Bolt Set","缸盖螺栓套件","HWA7-ENG-006",18.00,10,"OEM New",""),
        ("Engine & Fuel System","Fuel Tank 400L Alu","400L铝制燃油箱","HWA7-ENG-007",225.00,1,"OEM New",""),
        ("Engine & Fuel System","DEF/AdBlue Pump","尿素泵总成","HWA7-ENG-008",85.00,2,"OEM New","Euro V"),
        ("Engine & Fuel System","DEF Tank 55L","55L尿素箱","HWA7-ENG-009",48.00,3,"OEM New","With heater"),
        # Transmission & Drivetrain
        ("Transmission & Drivetrain","Gearbox HW19710 / Fast 9JS","HW19710变速箱","HWA7-TRN-001",1350.00,1,"OEM New",""),
        ("Transmission & Drivetrain","Clutch Disc 430mm","430mm离合器片","HWA7-TRN-002",48.00,5,"OEM New",""),
        ("Transmission & Drivetrain","Clutch Pressure Plate","离合器压盘","HWA7-TRN-003",72.00,3,"OEM New",""),
        ("Transmission & Drivetrain","Drive Shaft Assembly","传动轴总成","HWA7-TRN-004",165.00,2,"OEM New",""),
        ("Transmission & Drivetrain","Gearbox Oil Seal Kit","变速箱油封套件","HWA7-TRN-005",18.00,8,"OEM New",""),
        # Axle & Differential
        ("Axle & Differential","Front Axle HC16 Assy","HC16前桥总成","HWA7-AXL-001",720.00,1,"OEM New","16T"),
        ("Axle & Differential","Rear Drive Axle HC16","HC16驱动桥","HWA7-AXL-002",980.00,1,"OEM New",""),
        ("Axle & Differential","Front Hub & Bearing","前轮毂轴承总成","HWA7-AXL-003",68.00,4,"OEM New",""),
        ("Axle & Differential","King Pin Kit","主销套件","HWA7-AXL-004",18.00,10,"OEM New",""),
        ("Axle & Differential","5th Wheel Coupling","鞍座总成","HWA7-AXL-005",285.00,1,"OEM New","2-inch kingpin"),
        # Brakes & Air System
        ("Brakes & Air System","Front Brake Chamber T30","前制动气室","HWA7-BRK-001",28.00,8,"OEM New",""),
        ("Brakes & Air System","Rear Spring Brake T30/30","后弹簧制动气室","HWA7-BRK-002",45.00,6,"OEM New",""),
        ("Brakes & Air System","Brake Drum Front","前制动鼓","HWA7-BRK-003",52.00,4,"OEM New",""),
        ("Brakes & Air System","Brake Drum Rear","后制动鼓","HWA7-BRK-004",55.00,4,"OEM New",""),
        ("Brakes & Air System","Brake Lining Front Set","前制动蹄片","HWA7-BRK-005",22.00,8,"OEM New",""),
        ("Brakes & Air System","ABS Module ECU","ABS控制模块","HWA7-BRK-006",185.00,1,"OEM New","WABCO/KNORR"),
        ("Brakes & Air System","Air Dryer Cartridge","空气干燥器滤芯","HWA7-BRK-007",22.00,10,"OEM New",""),
        # Suspension & Steering
        ("Suspension & Steering","Front Leaf Spring","前板簧","HWA7-SUS-001",95.00,2,"OEM New","9-leaf"),
        ("Suspension & Steering","Air Suspension Bag Rear","后气囊悬挂","HWA7-SUS-002",85.00,4,"OEM New","Optional"),
        ("Suspension & Steering","Shock Absorber Front","前减振器","HWA7-SUS-003",38.00,4,"OEM New",""),
        ("Suspension & Steering","Steering Gear ZF8098","ZF8098转向机","HWA7-SUS-004",325.00,1,"OEM New",""),
        ("Suspension & Steering","Power Steering Pump","助力转向泵","HWA7-SUS-005",68.00,3,"OEM New",""),
        # Electrical & Lighting
        ("Electrical & Lighting","Left LED Headlight A7","A7左LED大灯","HWA7-ELC-001",65.00,3,"OEM New","LED type"),
        ("Electrical & Lighting","Right LED Headlight A7","A7右LED大灯","HWA7-ELC-002",65.00,3,"OEM New",""),
        ("Electrical & Lighting","LED Light Bar Front","前LED灯带","HWA7-ELC-003",38.00,4,"OEM New",""),
        ("Electrical & Lighting","Alternator 28V 70A","28V 70A发电机","HWA7-ELC-004",145.00,2,"OEM New",""),
        ("Electrical & Lighting","Starter Motor 24V","24V起动机","HWA7-ELC-005",115.00,2,"OEM New",""),
        ("Electrical & Lighting","Central ECU Controller","中央控制单元","HWA7-ELC-006",285.00,1,"OEM New",""),
        # Exhaust & Emission
        ("Exhaust & Emission","SCR Catalytic Converter","SCR催化转化器","HWA7-EXH-001",385.00,1,"OEM New","Euro V"),
        ("Exhaust & Emission","DPF Filter","DPF颗粒过滤器","HWA7-EXH-002",295.00,1,"OEM New",""),
        ("Exhaust & Emission","Exhaust Pipe Clamp Set","排气管夹套件","HWA7-EXH-003",18.00,10,"OEM New",""),
        # Cooling System
        ("Cooling System","Radiator Assembly","散热器总成","HWA7-COL-001",165.00,2,"OEM New","Al core"),
        ("Cooling System","Fan Clutch Viscous","粘性风扇离合器","HWA7-COL-002",95.00,2,"OEM New",""),
        ("Cooling System","Water Pump","水泵","HWA7-COL-003",48.00,4,"OEM New",""),
        ("Cooling System","Charge Air Cooler","增压中冷器","HWA7-COL-004",105.00,2,"OEM New",""),
        # Wheels & Tyres
        ("Wheels & Tyres","Alloy Wheel Rim 22.5in","22.5寸铝合金轮辋","HWA7-WHL-001",72.00,8,"OEM New","Forged"),
        ("Wheels & Tyres","Tyre 12R22.5 Drive","12R22.5驱动轮胎","HWA7-WHL-002",168.00,6,"OEM New",""),
        ("Wheels & Tyres","Tyre 315/80R22.5 Steer","315/80R22.5转向","HWA7-WHL-003",185.00,2,"OEM New",""),
        # Filters & Maintenance
        ("Filters & Maintenance","Engine Oil Filter","机油滤清器","HWA7-MNT-001",9.00,20,"OEM New",""),
        ("Filters & Maintenance","Fuel Filter Primary","初级油滤","HWA7-MNT-002",6.50,20,"OEM New",""),
        ("Filters & Maintenance","Air Filter","空气滤清器","HWA7-MNT-003",14.00,10,"OEM New",""),
        ("Filters & Maintenance","DEF/AdBlue Filter","尿素滤芯","HWA7-MNT-004",12.00,12,"OEM New",""),
        ("Filters & Maintenance","Cabin Air Filter","驾驶室空调滤","HWA7-MNT-005",7.00,15,"OEM New",""),
    ]
},

# ═══════════════════════════════════════════════════════════════════════════════
"SITRAK C7H": {
    "info": {
        "year": "2016–2024", "type": "Premium Heavy Tractor 4×2 / 6×4",
        "engine": "MAN D2676 / MC13 400–540HP", "gvw": "GVW 18–25T"
    },
    "parts": [
        # Cabin & Body
        ("Cabin & Body","Left Door C7H","C7H左车门总成","C7H-CAB-001",175.00,2,"OEM New","MAN license"),
        ("Cabin & Body","Right Door C7H","C7H右车门总成","C7H-CAB-002",175.00,2,"OEM New",""),
        ("Cabin & Body","Front Bumper C7H","C7H前保险杠","C7H-CAB-003",115.00,2,"OEM New","With LED DRL"),
        ("Cabin & Body","Windshield C7H","C7H前挡风玻璃","C7H-CAB-004",88.00,2,"OEM New","Heated"),
        ("Cabin & Body","Roof Spoiler C7H","顶置导流罩","C7H-CAB-005",185.00,1,"OEM New",""),
        ("Cabin & Body","Left Side Deflector","左侧导流板","C7H-CAB-006",55.00,3,"OEM New",""),
        ("Cabin & Body","Right Side Deflector","右侧导流板","C7H-CAB-007",55.00,3,"OEM New",""),
        ("Cabin & Body","Left Mirror C7H Electric","C7H左电动后视镜","C7H-CAB-008",62.00,2,"OEM New","Camera option"),
        ("Cabin & Body","Right Mirror C7H Electric","C7H右电动后视镜","C7H-CAB-009",62.00,2,"OEM New",""),
        ("Cabin & Body","Sleeping Berth Full","全卧铺总成","C7H-CAB-010",225.00,1,"OEM New","Luxury trim"),
        ("Cabin & Body","Foot Step Assy","脚踏板总成","C7H-CAB-011",38.00,4,"OEM New","Pair"),
        # Engine & Fuel System
        ("Engine & Fuel System","MAN D2676 Injector","MAN D2676喷油嘴","C7H-ENG-001",95.00,6,"OEM New","Bosch CRIN3"),
        ("Engine & Fuel System","Common Rail HP Pump","高压共轨油泵","C7H-ENG-002",325.00,1,"OEM New","Bosch CP4"),
        ("Engine & Fuel System","Turbocharger BorgWarner","BW涡轮增压器","C7H-ENG-003",285.00,1,"OEM New","VTG"),
        ("Engine & Fuel System","Fuel Tank 500L Alu","500L铝合金燃油箱","C7H-ENG-004",285.00,1,"OEM New",""),
        ("Engine & Fuel System","EGR Module Assy","EGR模块总成","C7H-ENG-005",185.00,1,"OEM New",""),
        ("Engine & Fuel System","AdBlue Injector Nozzle","尿素喷嘴","C7H-ENG-006",45.00,5,"OEM New",""),
        ("Engine & Fuel System","Air Intake Manifold","进气歧管","C7H-ENG-007",65.00,3,"OEM New",""),
        # Transmission & Drivetrain
        ("Transmission & Drivetrain","ZF AMT 12TX Gearbox","ZF 12TX自动变速箱","C7H-TRN-001",3850.00,1,"OEM New","ZF TraXon/AS-Tronic"),
        ("Transmission & Drivetrain","Clutch Kit 430mm","430mm离合器套件","C7H-TRN-002",125.00,3,"OEM New","Sachs brand"),
        ("Transmission & Drivetrain","Drive Shaft CV","等速传动轴","C7H-TRN-003",185.00,2,"OEM New",""),
        ("Transmission & Drivetrain","Retarder Voith","Voith液力缓速器","C7H-TRN-004",1850.00,1,"OEM New","Optional"),
        # Axle & Differential
        ("Axle & Differential","Front Axle MAN Type","MAN型前桥总成","C7H-AXL-001",850.00,1,"OEM New","16T"),
        ("Axle & Differential","Drive Axle MAN Hyp.","MAN双曲线驱动桥","C7H-AXL-002",1150.00,1,"OEM New","13T"),
        ("Axle & Differential","5th Wheel 3.5in","3.5寸鞍座","C7H-AXL-003",325.00,1,"OEM New","JOST brand"),
        ("Axle & Differential","King Pin Kit MAN","MAN主销套件","C7H-AXL-004",25.00,8,"OEM New",""),
        # Brakes & Air System
        ("Brakes & Air System","WABCO ABS ECU","WABCO ABS模块","C7H-BRK-001",285.00,1,"OEM New",""),
        ("Brakes & Air System","EBS Brake Control","EBS电子制动系统","C7H-BRK-002",485.00,1,"OEM New","WABCO"),
        ("Brakes & Air System","Disc Brake Caliper F","前盘式制动卡钳","C7H-BRK-003",185.00,2,"OEM New","KNORR"),
        ("Brakes & Air System","Brake Disc Front","前制动盘","C7H-BRK-004",95.00,4,"OEM New","Vented"),
        ("Brakes & Air System","Brake Disc Rear","后制动盘","C7H-BRK-005",88.00,4,"OEM New",""),
        ("Brakes & Air System","Brake Pad Set Front","前刹车片套件","C7H-BRK-006",38.00,8,"OEM New",""),
        ("Brakes & Air System","Air Compressor MAN","MAN型空压机","C7H-BRK-007",185.00,2,"OEM New",""),
        ("Brakes & Air System","Air Dryer WABCO","WABCO空气干燥器","C7H-BRK-008",125.00,2,"OEM New",""),
        # Suspension & Steering
        ("Suspension & Steering","Front Parabolic Spring","前抛物线板簧","C7H-SUS-001",125.00,2,"OEM New","2-leaf"),
        ("Suspension & Steering","Rear Air Bag Assy","后气囊总成","C7H-SUS-002",95.00,4,"OEM New","MAN spec"),
        ("Suspension & Steering","Shock Absorber Front","前减振器","C7H-SUS-003",65.00,4,"OEM New","KONI/Sachs"),
        ("Suspension & Steering","ZF Power Steering Unit","ZF助力转向总成","C7H-SUS-004",425.00,1,"OEM New","ZF 8099"),
        ("Suspension & Steering","Track Rod Assembly","转向横拉杆","C7H-SUS-005",55.00,3,"OEM New",""),
        # Electrical & Lighting
        ("Electrical & Lighting","Left Full LED Headlight","左全LED大灯","C7H-ELC-001",95.00,2,"OEM New","DRL+main+high"),
        ("Electrical & Lighting","Right Full LED Headlight","右全LED大灯","C7H-ELC-002",95.00,2,"OEM New",""),
        ("Electrical & Lighting","LED Marker Light Set","LED示廓灯套件","C7H-ELC-003",28.00,6,"OEM New","7pcs"),
        ("Electrical & Lighting","Alternator 28V 80A","28V 80A发电机","C7H-ELC-004",165.00,2,"OEM New",""),
        ("Electrical & Lighting","Central Gateway ECU","中央网关ECU","C7H-ELC-005",385.00,1,"OEM New","CAN bus"),
        ("Electrical & Lighting","Instrument Cluster LCD","LCD组合仪表","C7H-ELC-006",225.00,1,"OEM New","7in color"),
        # Exhaust & Emission
        ("Exhaust & Emission","SCR System Complete","SCR系统总成","C7H-EXH-001",485.00,1,"OEM New","Euro V/VI"),
        ("Exhaust & Emission","DPF Regen System","DPF再生系统","C7H-EXH-002",385.00,1,"OEM New",""),
        ("Exhaust & Emission","Exhaust Brake Valve","排气制动阀","C7H-EXH-003",65.00,4,"OEM New",""),
        # Cooling System
        ("Cooling System","Radiator Module Assy","散热模块总成","C7H-COL-001",225.00,1,"OEM New","Rad+intercooler+fan"),
        ("Cooling System","Viscous Fan Drive","粘性风扇离合器","C7H-COL-002",115.00,2,"OEM New",""),
        ("Cooling System","Water Pump MAN","MAN水泵","C7H-COL-003",65.00,3,"OEM New",""),
        ("Cooling System","Charge Air Cooler","中冷器","C7H-COL-004",125.00,2,"OEM New",""),
        # Wheels & Tyres
        ("Wheels & Tyres","Alloy Wheel 22.5in Super","22.5寸超轻铝轮辋","C7H-WHL-001",88.00,8,"OEM New","Forged alcoa type"),
        ("Wheels & Tyres","Tyre 315/70R22.5 Drive","315/70R22.5驱动胎","C7H-WHL-002",195.00,6,"OEM New",""),
        ("Wheels & Tyres","Tyre 315/70R22.5 Steer","315/70R22.5转向胎","C7H-WHL-003",185.00,4,"OEM New",""),
        # Filters & Maintenance
        ("Filters & Maintenance","Engine Oil Filter MAN","MAN机油滤清器","C7H-MNT-001",14.00,15,"OEM New","Spin-on"),
        ("Filters & Maintenance","Fuel Filter MAN","MAN燃油滤清器","C7H-MNT-002",12.00,15,"OEM New",""),
        ("Filters & Maintenance","Air Filter Primary","空气滤清器","C7H-MNT-003",18.00,8,"OEM New",""),
        ("Filters & Maintenance","AdBlue Filter","尿素滤芯","C7H-MNT-004",15.00,10,"OEM New",""),
        ("Filters & Maintenance","Gearbox Oil ZF","ZF变速箱油","C7H-MNT-005",28.00,8,"OEM New","5L ZF-Lifeguard"),
    ]
},

# ═══════════════════════════════════════════════════════════════════════════════
"SITRAK G7": {
    "info": {
        "year": "2019–2024", "type": "Premium Heavy Tractor 4×2 / 6×4",
        "engine": "MC13 430–560HP Euro VI", "gvw": "GVW 18–25T"
    },
    "parts": [
        # Cabin & Body
        ("Cabin & Body","Left Door G7","G7左车门总成","SG7-CAB-001",185.00,2,"OEM New","G7 series"),
        ("Cabin & Body","Right Door G7","G7右车门总成","SG7-CAB-002",185.00,2,"OEM New",""),
        ("Cabin & Body","Front Bumper G7 LED","G7前保险杠LED版","SG7-CAB-003",128.00,2,"OEM New",""),
        ("Cabin & Body","Windshield G7 Heated","G7电热前挡风玻璃","SG7-CAB-004",95.00,2,"OEM New",""),
        ("Cabin & Body","Roof Air Deflector G7","G7顶部导流罩","SG7-CAB-005",195.00,1,"OEM New","Aero optimized"),
        ("Cabin & Body","Left Mirror Camera G7","G7左摄像头后视镜","SG7-CAB-006",78.00,2,"OEM New","HD camera"),
        ("Cabin & Body","Right Mirror Camera G7","G7右摄像头后视镜","SG7-CAB-007",78.00,2,"OEM New",""),
        ("Cabin & Body","Cabin Air Suspension","驾驶室空气悬挂","SG7-CAB-008",385.00,1,"OEM New","Full air"),
        ("Cabin & Body","Luxury Seat Driver","豪华驾驶座椅","SG7-CAB-009",285.00,1,"OEM New","Air susp+heat+cool"),
        ("Cabin & Body","Cabin Step Set","驾驶室踏步套件","SG7-CAB-010",42.00,4,"OEM New",""),
        # Engine & Fuel System
        ("Engine & Fuel System","MC13 Common Rail Injector","MC13共轨喷射器","SG7-ENG-001",98.00,6,"OEM New","Bosch CRIN4"),
        ("Engine & Fuel System","HP Fuel Pump Bosch","Bosch高压燃油泵","SG7-ENG-002",365.00,1,"OEM New","CP4.2"),
        ("Engine & Fuel System","Turbo BorgWarner VTG","BW VTG涡轮","SG7-ENG-003",325.00,1,"OEM New",""),
        ("Engine & Fuel System","Fuel Tank 600L","600L铝制油箱","SG7-ENG-004",325.00,1,"OEM New",""),
        ("Engine & Fuel System","SCR Injector Doser","SCR尿素喷射器","SG7-ENG-005",55.00,5,"OEM New",""),
        ("Engine & Fuel System","AdBlue Tank 80L","80L尿素箱","SG7-ENG-006",65.00,3,"OEM New","With pump"),
        ("Engine & Fuel System","EGR Cooler G7","EGR冷却器","SG7-ENG-007",95.00,2,"OEM New",""),
        ("Engine & Fuel System","Engine ECU","发动机控制单元","SG7-ENG-008",485.00,1,"OEM New","Prog. required"),
        # Transmission & Drivetrain
        ("Transmission & Drivetrain","ZF AMT TraXon 16 Spd","ZF AMT 16速自动箱","SG7-TRN-001",4250.00,1,"OEM New","TraXon series"),
        ("Transmission & Drivetrain","Clutch Actuator AMT","AMT离合器执行器","SG7-TRN-002",285.00,1,"OEM New",""),
        ("Transmission & Drivetrain","Clutch Disc Sachs","Sachs离合器片","SG7-TRN-003",85.00,3,"OEM New","430mm"),
        ("Transmission & Drivetrain","Drive Shaft Assy","传动轴总成","SG7-TRN-004",195.00,2,"OEM New",""),
        ("Transmission & Drivetrain","Retarder Voith 115","Voith缓速器115","SG7-TRN-005",2250.00,1,"OEM New","Optional"),
        # Axle & Differential
        ("Axle & Differential","Front Axle 16T Disc","16T前桥盘式制动","SG7-AXL-001",925.00,1,"OEM New",""),
        ("Axle & Differential","Rear Tandem Axle 13T","13T后桥总成","SG7-AXL-002",1250.00,1,"OEM New",""),
        ("Axle & Differential","5th Wheel JOST W24","JOST W24鞍座","SG7-AXL-003",385.00,1,"OEM New","35T rating"),
        ("Axle & Differential","Lift Axle System","提升轴系统","SG7-AXL-004",1250.00,1,"OEM New","Optional 3rd axle"),
        # Brakes & Air System
        ("Brakes & Air System","EBS System WABCO","WABCO EBS系统","SG7-BRK-001",625.00,1,"OEM New","Euro VI"),
        ("Brakes & Air System","AEBS Emergency Brake","AEBS自动紧急制动","SG7-BRK-002",1250.00,1,"OEM New","Radar sensor incl."),
        ("Brakes & Air System","Disc Brake Caliper","盘式制动卡钳","SG7-BRK-003",195.00,2,"OEM New","KNORR SB"),
        ("Brakes & Air System","Brake Disc 430mm","430mm制动盘","SG7-BRK-004",105.00,4,"OEM New",""),
        ("Brakes & Air System","Brake Pad Set","刹车片套件","SG7-BRK-005",42.00,8,"OEM New","Low metallic"),
        ("Brakes & Air System","Lane Departure LDWS","车道偏离预警系统","SG7-BRK-006",485.00,1,"OEM New","Camera based"),
        # Suspension & Steering
        ("Suspension & Steering","Full Air Suspension Kit","全气囊悬挂套件","SG7-SUS-001",285.00,1,"OEM New","Front+rear"),
        ("Suspension & Steering","Air Bag Front","前气囊","SG7-SUS-002",95.00,4,"OEM New",""),
        ("Suspension & Steering","Air Bag Rear","后气囊","SG7-SUS-003",88.00,4,"OEM New",""),
        ("Suspension & Steering","Levelling Valve","调平阀","SG7-SUS-004",45.00,4,"OEM New","WABCO"),
        ("Suspension & Steering","ZF Active Steering","ZF主动转向系统","SG7-SUS-005",1850.00,1,"OEM New","Rear-axle steering"),
        ("Suspension & Steering","Shock Absorber KONI","KONI减振器","SG7-SUS-006",72.00,4,"OEM New",""),
        # Electrical & Lighting
        ("Electrical & Lighting","Full LED Headlight G7","G7全LED大灯组","SG7-ELC-001",125.00,2,"OEM New","Matrix LED"),
        ("Electrical & Lighting","LED Fog Light G7","G7 LED雾灯","SG7-ELC-002",38.00,4,"OEM New",""),
        ("Electrical & Lighting","TPMS System","胎压监测系统","SG7-ELC-003",185.00,1,"OEM New","6-wheel"),
        ("Electrical & Lighting","Telematics Unit 4G","4G车载终端","SG7-ELC-004",125.00,1,"OEM New","GPS+Beidou"),
        ("Electrical & Lighting","Alternator 28V 100A","28V 100A发电机","SG7-ELC-005",195.00,2,"OEM New",""),
        ("Electrical & Lighting","Color Instrument Cluster","彩色仪表盘","SG7-ELC-006",285.00,1,"OEM New","10in display"),
        # Exhaust & Emission
        ("Exhaust & Emission","SCR+DPF Combined","SCR+DPF组合器","SG7-EXH-001",625.00,1,"OEM New","Euro VI"),
        ("Exhaust & Emission","DOC Oxidation Catalyst","DOC氧化催化器","SG7-EXH-002",285.00,1,"OEM New",""),
        ("Exhaust & Emission","Exhaust Aftertreatment","后处理系统","SG7-EXH-003",925.00,1,"OEM New","Complete kit"),
        # Cooling System
        ("Cooling System","Cooling Module Assy","冷却模块总成","SG7-COL-001",265.00,1,"OEM New",""),
        ("Cooling System","Electric Fan Control","电控风扇总成","SG7-COL-002",145.00,2,"OEM New",""),
        ("Cooling System","Water Pump","水泵总成","SG7-COL-003",68.00,3,"OEM New",""),
        # Wheels & Tyres
        ("Wheels & Tyres","Forged Alloy Wheel 22.5","22.5寸锻造铝轮","SG7-WHL-001",95.00,8,"OEM New","Super single opt."),
        ("Wheels & Tyres","Super Single Tyre 385/65R22.5","385/65R22.5超宽单胎","SG7-WHL-002",228.00,4,"OEM New",""),
        ("Wheels & Tyres","Standard Tyre 315/70R22.5","315/70R22.5标准轮","SG7-WHL-003",195.00,6,"OEM New",""),
        # Filters & Maintenance
        ("Filters & Maintenance","Full Service Kit MC13","MC13全套保养包","SG7-MNT-001",65.00,5,"OEM New","Oil+fuel+air filter"),
        ("Filters & Maintenance","Gearbox Oil ZF S671","ZF S671变速箱油","SG7-MNT-002",32.00,8,"OEM New","5L"),
        ("Filters & Maintenance","Axle Oil 90W","驱动桥油90W","SG7-MNT-003",22.00,10,"OEM New","3L GL-5"),
        ("Filters & Maintenance","AdBlue Fluid 10L","10L尿素液","SG7-MNT-004",8.00,20,"OEM New","AUS32 ISO22241"),
    ]
},

# ═══════════════════════════════════════════════════════════════════════════════
"HOWO Light Truck (轻卡)": {
    "info": {
        "year": "2015–2024", "type": "Light Cargo Truck 4×2",
        "engine": "WD618 180–240HP", "gvw": "GVW 6–12T"
    },
    "parts": [
        # Cabin & Body
        ("Cabin & Body","Left Door Light Truck","轻卡左车门","HWL-CAB-001",85.00,3,"OEM New",""),
        ("Cabin & Body","Right Door Light Truck","轻卡右车门","HWL-CAB-002",85.00,3,"OEM New",""),
        ("Cabin & Body","Front Bumper","前保险杠","HWL-CAB-003",48.00,4,"OEM New",""),
        ("Cabin & Body","Windshield","前挡风玻璃","HWL-CAB-004",52.00,3,"OEM New",""),
        ("Cabin & Body","Engine Hood","发动机盖","HWL-CAB-005",65.00,3,"OEM New",""),
        ("Cabin & Body","Left Side Mirror","左后视镜","HWL-CAB-006",22.00,6,"OEM New",""),
        ("Cabin & Body","Right Side Mirror","右后视镜","HWL-CAB-007",22.00,6,"OEM New",""),
        ("Cabin & Body","Cargo Body 5.2m Steel","5.2米钢货箱","HWL-CAB-008",1250.00,1,"OEM New","Side open type"),
        ("Cabin & Body","Tailgate Cargo Body","货箱后板","HWL-CAB-009",88.00,2,"OEM New",""),
        # Engine & Fuel System
        ("Engine & Fuel System","Fuel Injector WD618","WD618喷油嘴","HWL-ENG-001",32.00,6,"OEM New",""),
        ("Engine & Fuel System","Fuel Pump WD618","WD618燃油泵","HWL-ENG-002",85.00,2,"OEM New",""),
        ("Engine & Fuel System","Turbocharger","涡轮增压器","HWL-ENG-003",135.00,2,"OEM New",""),
        ("Engine & Fuel System","Engine Gasket Set","发动机衬垫套件","HWL-ENG-004",22.00,6,"OEM New","Full set"),
        ("Engine & Fuel System","Fuel Tank 150L","150L燃油箱","HWL-ENG-005",95.00,2,"OEM New","Steel"),
        # Transmission & Drivetrain
        ("Transmission & Drivetrain","Gearbox 6-Speed","6速变速箱","HWL-TRN-001",485.00,1,"OEM New",""),
        ("Transmission & Drivetrain","Clutch Kit 350mm","350mm离合器套件","HWL-TRN-002",62.00,4,"OEM New","3pcs"),
        ("Transmission & Drivetrain","Drive Shaft","传动轴","HWL-TRN-003",88.00,3,"OEM New",""),
        # Axle & Differential
        ("Axle & Differential","Front Axle 6T","6T前桥总成","HWL-AXL-001",325.00,1,"OEM New",""),
        ("Axle & Differential","Rear Axle 10T","10T后桥总成","HWL-AXL-002",485.00,1,"OEM New",""),
        ("Axle & Differential","Wheel Hub Front","前轮毂","HWL-AXL-003",42.00,4,"OEM New",""),
        ("Axle & Differential","Wheel Hub Rear","后轮毂","HWL-AXL-004",48.00,4,"OEM New",""),
        # Brakes & Air System
        ("Brakes & Air System","Front Brake Drum","前制动鼓","HWL-BRK-001",38.00,4,"OEM New",""),
        ("Brakes & Air System","Rear Brake Drum","后制动鼓","HWL-BRK-002",42.00,4,"OEM New",""),
        ("Brakes & Air System","Brake Shoe Set","刹车蹄片套件","HWL-BRK-003",18.00,8,"OEM New",""),
        ("Brakes & Air System","Brake Master Cylinder","制动主缸","HWL-BRK-004",38.00,4,"OEM New","Hydraulic type"),
        ("Brakes & Air System","Air Compressor","空压机","HWL-BRK-005",85.00,3,"OEM New","Air brake model"),
        # Suspension & Steering
        ("Suspension & Steering","Front Leaf Spring","前板簧","HWL-SUS-001",62.00,3,"OEM New",""),
        ("Suspension & Steering","Rear Leaf Spring","后板簧","HWL-SUS-002",72.00,3,"OEM New",""),
        ("Suspension & Steering","Shock Absorber Front","前减振器","HWL-SUS-003",25.00,6,"OEM New",""),
        ("Suspension & Steering","Shock Absorber Rear","后减振器","HWL-SUS-004",22.00,6,"OEM New",""),
        ("Suspension & Steering","Steering Gear Box","转向器","HWL-SUS-005",145.00,2,"OEM New",""),
        ("Suspension & Steering","Tie Rod End","横拉杆球头","HWL-SUS-006",12.00,10,"OEM New",""),
        # Electrical & Lighting
        ("Electrical & Lighting","Left Headlight","左大灯","HWL-ELC-001",28.00,5,"OEM New",""),
        ("Electrical & Lighting","Right Headlight","右大灯","HWL-ELC-002",28.00,5,"OEM New",""),
        ("Electrical & Lighting","Tail Light Left","左后尾灯","HWL-ELC-003",18.00,6,"OEM New",""),
        ("Electrical & Lighting","Tail Light Right","右后尾灯","HWL-ELC-004",18.00,6,"OEM New",""),
        ("Electrical & Lighting","Alternator 24V 45A","24V 45A发电机","HWL-ELC-005",85.00,3,"OEM New",""),
        ("Electrical & Lighting","Starter Motor 24V","24V起动机","HWL-ELC-006",75.00,3,"OEM New",""),
        # Cooling System
        ("Cooling System","Radiator Assembly","散热器总成","HWL-COL-001",88.00,3,"OEM New",""),
        ("Cooling System","Water Pump","水泵","HWL-COL-002",32.00,5,"OEM New",""),
        ("Cooling System","Fan Belt Set","皮带套件","HWL-COL-003",12.00,15,"OEM New",""),
        # Wheels & Tyres
        ("Wheels & Tyres","Steel Wheel 17.5in","17.5寸钢圈","HWL-WHL-001",28.00,8,"OEM New","8×17.5"),
        ("Wheels & Tyres","Tyre 7.50R16 Light","7.50R16轻卡轮胎","HWL-WHL-002",62.00,8,"OEM New",""),
        ("Wheels & Tyres","Tyre 8.25R16","8.25R16轮胎","HWL-WHL-003",72.00,8,"OEM New",""),
        # Filters & Maintenance
        ("Filters & Maintenance","Engine Oil Filter","机油滤清器","HWL-MNT-001",6.50,20,"OEM New",""),
        ("Filters & Maintenance","Fuel Filter","燃油滤清器","HWL-MNT-002",5.00,20,"OEM New",""),
        ("Filters & Maintenance","Air Filter","空气滤清器","HWL-MNT-003",9.00,12,"OEM New",""),
        ("Filters & Maintenance","Engine Oil 15W-40 5L","15W-40机油5升","HWL-MNT-004",15.00,12,"OEM New","CI-4"),
        ("Filters & Maintenance","Wiper Blades","雨刮器","HWL-MNT-005",8.00,10,"OEM New","Pair"),
    ]
},

} # end MODELS

# ─── Excel builder (identical structure to BYD script) ───────────────────────

def make_fill(hex_color):
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type='solid')

def make_border():
    s = Side(style='thin', color='CCCCCC')
    return Border(left=s, right=s, top=s, bottom=s)

def make_font(bold=False, size=10, color='000000', name='Calibri'):
    return Font(bold=bold, size=size, color=color, name=name)

HEADERS    = ['#','Category','Part Name (EN)','Part Name (CN)','Part Number',
              'Unit Price (USD)','MOQ (pcs)','Condition','Notes']
COL_WIDTHS = [5, 24, 34, 24, 18, 16, 11, 14, 28]

def write_model_sheet(wb, model_name, model_data):
    safe = model_name.replace(':','').replace('/','×').replace(' ','_')[:31]
    ws = wb.create_sheet(title=safe)

    # Row 1 — title banner
    ws.merge_cells('A1:I1')
    ws['A1'] = model_name + '  ·  Sinotruk Spare Parts  |  CM IMPORT&EXPORT'
    ws['A1'].fill      = make_fill(C_DARK)
    ws['A1'].font      = Font(bold=True, size=14, color=C_WHITE, name='Calibri')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    # Row 2 — model info
    info = model_data['info']
    info_txt = (f"  Year: {info.get('year','')}   |   Type: {info.get('type','')}"
                f"   |   Engine: {info.get('engine','')}   |   {info.get('gvw','')}"
                f"   |   FOB China — All prices USD")
    ws.merge_cells('A2:I2')
    ws['A2'] = info_txt
    ws['A2'].fill      = make_fill(C_ACCENT)
    ws['A2'].font      = Font(italic=True, size=9, color='D0E8FF', name='Calibri')
    ws['A2'].alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[2].height = 18

    # Row 3 — column headers
    for ci, (hdr, w) in enumerate(zip(HEADERS, COL_WIDTHS), start=1):
        c = ws.cell(row=3, column=ci, value=hdr)
        c.fill      = make_fill(C_RED_MID)
        c.font      = Font(bold=True, size=10, color=C_WHITE, name='Calibri')
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border    = make_border()
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[3].height = 22

    # Data rows
    parts    = model_data['parts']
    prev_cat = None
    row_num  = 4
    idx      = 1

    for part in parts:
        cat, name_en, name_cn, part_no, price, moq, cond, notes = part

        # Category separator
        if cat != prev_cat:
            cat_color = CAT_COLORS.get(cat, "555555")
            ws.merge_cells(f'A{row_num}:I{row_num}')
            c = ws.cell(row=row_num, column=1,
                        value=f'  {cat}')
            c.fill      = make_fill(cat_color)
            c.font      = Font(bold=True, size=10, color=C_WHITE, name='Calibri')
            c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
            ws.row_dimensions[row_num].height = 18
            row_num  += 1
            prev_cat  = cat

        row_fill = make_fill('FFFFFF') if idx % 2 == 1 else make_fill(C_LIGHT_GRAY)
        values = [idx, cat, name_en, name_cn, part_no, price, moq, cond, notes]

        for ci, val in enumerate(values, start=1):
            c = ws.cell(row=row_num, column=ci, value=val)
            c.fill   = row_fill
            c.border = make_border()
            c.font   = make_font(size=9)
            if ci == 1:
                c.alignment = Alignment(horizontal='center', vertical='center')
            elif ci == 6:
                c.number_format = '"$"#,##0.00'
                c.alignment     = Alignment(horizontal='right', vertical='center')
                c.font          = Font(bold=True, size=9, color='1A6B2A', name='Calibri')
            elif ci == 7:
                c.alignment = Alignment(horizontal='center', vertical='center')
            else:
                c.alignment = Alignment(horizontal='left', vertical='center',
                                        wrap_text=(ci in [3,4,9]))
        ws.row_dimensions[row_num].height = 16
        row_num += 1
        idx     += 1

    ws.freeze_panes        = 'A4'
    ws.auto_filter.ref     = f'A3:I{row_num-1}'
    return ws


def write_index_sheet(wb, models_dict):
    ws = wb.create_sheet(title='INDEX', index=0)

    # Title
    ws.merge_cells('A1:G1')
    ws['A1'] = 'Sinotruk (中国重汽) Spare Parts Catalog — By Vehicle Model'
    ws['A1'].fill      = make_fill(C_RED_MID)
    ws['A1'].font      = Font(bold=True, size=15, color=C_WHITE, name='Calibri')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 36

    ws.merge_cells('A2:G2')
    ws['A2'] = 'CM IMPORT&EXPORT  |  FOB China  |  All prices USD  |  Wholesale  |  HOWO · SITRAK · A7 · Light Truck'
    ws['A2'].fill      = make_fill(C_ACCENT)
    ws['A2'].font      = Font(italic=True, size=10, color='D0E8FF', name='Calibri')
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 20

    idx_hdrs   = ['#','Model','Type','Year','Engine / Power','Total Parts','Sheet Link']
    idx_widths = [5, 28, 30, 14, 30, 13, 26]
    for ci, (h, w) in enumerate(zip(idx_hdrs, idx_widths), start=1):
        c = ws.cell(row=3, column=ci, value=h)
        c.fill      = make_fill(C_DARK)
        c.font      = Font(bold=True, size=11, color=C_WHITE, name='Calibri')
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border    = make_border()
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[3].height = 24

    for i, (mname, mdata) in enumerate(models_dict.items(), start=1):
        row   = 3 + i
        info  = mdata['info']
        safe  = mname.replace(':','').replace('/','×').replace(' ','_')[:31]
        total = len(mdata['parts'])
        vals  = [i, mname, info.get('type',''), info.get('year',''),
                 info.get('engine',''), total, f'See sheet: {safe}']
        rfill = make_fill('FFFFFF') if i % 2 == 1 else make_fill(C_LIGHT_GRAY)
        for ci, v in enumerate(vals, start=1):
            c = ws.cell(row=row, column=ci, value=v)
            c.fill      = rfill
            c.border    = make_border()
            c.font      = make_font(size=10)
            c.alignment = Alignment(
                horizontal='center' if ci in [1,6] else 'left',
                vertical='center')
        ws.row_dimensions[row].height = 20

    total_row = 3 + len(models_dict) + 1
    ws.cell(row=total_row, column=1, value='TOTAL')
    ws.cell(row=total_row, column=6, value=sum(len(v['parts']) for v in models_dict.values()))
    for ci in range(1, 8):
        c = ws.cell(row=total_row, column=ci)
        c.fill      = make_fill(C_RED_MID)
        c.font      = Font(bold=True, size=11, color=C_WHITE, name='Calibri')
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border    = make_border()
    ws.row_dimensions[total_row].height = 22
    return ws


def build_workbook():
    wb = openpyxl.Workbook()
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    print('Building INDEX sheet ...')
    write_index_sheet(wb, MODELS)

    for model_name, model_data in MODELS.items():
        print(f'  Writing: {model_name} ({len(model_data["parts"])} parts) ...')
        write_model_sheet(wb, model_name, model_data)

    wb.save(OUTPUT_FILE)
    total = sum(len(v['parts']) for v in MODELS.values())
    print(f'\n[OK] Saved -> {OUTPUT_FILE}')
    print(f'[OK] {len(MODELS)} model sheets + 1 INDEX = {len(MODELS)+1} sheets')
    print(f'[OK] {total} parts total')


if __name__ == '__main__':
    build_workbook()
